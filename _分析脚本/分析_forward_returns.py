import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook('算展D.sh600155.xlsx', data_only=True)
ws = wb['LD600155']

all_rows = list(range(2, ws.max_row + 1))

# Build forward targets
forward_chg_1 = {}
forward_chg_5 = {}
forward_chg_20 = {}

for i, row in enumerate(all_rows):
    chg = ws.cell(row, 5).value
    if chg is None:
        continue

    if i + 1 < len(all_rows):
        r_next = all_rows[i + 1]
        next_chg = ws.cell(r_next, 5).value
        if next_chg is not None:
            forward_chg_1[row] = next_chg

    if i + 5 < len(all_rows):
        s = 0
        for j in range(1, 6):
            c = ws.cell(all_rows[i + j], 5).value
            if c is not None:
                s += c
        forward_chg_5[row] = s

    if i + 20 < len(all_rows):
        s = 0
        for j in range(1, 21):
            c = ws.cell(all_rows[i + j], 5).value
            if c is not None:
                s += c
        forward_chg_20[row] = s

print("=== 1. 仓周分类 vs 下日涨跌 ===")
groups = defaultdict(list)
for row in all_rows:
    if row in forward_chg_1:
        cw = ws.cell(row, 225).value
        if cw:
            groups[str(cw)].append(forward_chg_1[row])

for k, vals in sorted(groups.items()):
    if len(vals) >= 5:
        pos = sum(1 for v in vals if v > 0)
        avg = sum(vals) / len(vals)
        print(f"  仓周={k:>5}: 样本={len(vals):>5} 涨概率={pos/len(vals)*100:>5.1f}% 平均={avg:>7.2f}")

print("\n=== 2. 仓日分类 vs 下日涨跌 ===")
groups2 = defaultdict(list)
for row in all_rows:
    if row in forward_chg_1:
        cd = ws.cell(row, 226).value
        if cd:
            groups2[str(cd)].append(forward_chg_1[row])

for k, vals in sorted(groups2.items()):
    if len(vals) >= 5:
        pos = sum(1 for v in vals if v > 0)
        avg = sum(vals) / len(vals)
        print(f"  仓日={k:>5}: 样本={len(vals):>5} 涨概率={pos/len(vals)*100:>5.1f}% 平均={avg:>7.2f}")

print("\n=== 3. 大局(首字) vs 下周涨跌 ===")
groups3 = defaultdict(list)
for row in all_rows:
    if row in forward_chg_5:
        bp = ws.cell(row, 88).value
        if bp:
            key = str(bp)[:2]
            groups3[key].append(forward_chg_5[row])

for k, vals in sorted(groups3.items()):
    if len(vals) >= 5:
        pos = sum(1 for v in vals if v > 0)
        avg = sum(vals) / len(vals)
        print(f"  大局={k:>4}: 样本={len(vals):>5} 涨概率={pos/len(vals)*100:>5.1f}% 平均={avg:>7.2f}")

print("\n=== 4. 波类周十 vs 下周涨跌 ===")
groups4 = defaultdict(list)
for row in all_rows:
    if row in forward_chg_5:
        wt = ws.cell(row, 50).value
        if wt:
            groups4[str(wt)].append(forward_chg_5[row])

for k, vals in sorted(groups4.items()):
    if len(vals) >= 5:
        pos = sum(1 for v in vals if v > 0)
        avg = sum(vals) / len(vals)
        print(f"  波类={k:>10}: 样本={len(vals):>5} 涨概率={pos/len(vals)*100:>5.1f}% 平均={avg:>7.2f}")

print("\n=== 5. 护型WXAB周 vs 下周涨跌 ===")
groups5 = defaultdict(list)
for row in all_rows:
    if row in forward_chg_5:
        pt = ws.cell(row, 79).value
        if pt:
            groups5[str(pt)].append(forward_chg_5[row])

for k, vals in sorted(groups5.items()):
    if len(vals) >= 5:
        pos = sum(1 for v in vals if v > 0)
        avg = sum(vals) / len(vals)
        print(f"  护型={k:>10}: 样本={len(vals):>5} 涨概率={pos/len(vals)*100:>5.1f}% 平均={avg:>7.2f}")

print("\n=== 6. 冲否周十 vs 下周涨跌 ===")
groups6 = defaultdict(list)
for row in all_rows:
    if row in forward_chg_5:
        ch = ws.cell(row, 51).value
        if ch:
            groups6[str(ch)].append(forward_chg_5[row])

for k, vals in sorted(groups6.items()):
    if len(vals) >= 5:
        pos = sum(1 for v in vals if v > 0)
        avg = sum(vals) / len(vals)
        print(f"  冲否={k:>5}: 样本={len(vals):>5} 涨概率={pos/len(vals)*100:>5.1f}% 平均={avg:>7.2f}")

print("\n=== 7. 盈提示 vs 下日涨跌 ===")
groups7 = defaultdict(list)
for row in all_rows:
    if row in forward_chg_1:
        ph = ws.cell(row, 92).value  # d_profit_hint = col CN=92
        if ph:
            groups7[str(ph)].append(forward_chg_1[row])

for k, vals in sorted(groups7.items()):
    if len(vals) >= 3:
        pos = sum(1 for v in vals if v > 0)
        avg = sum(vals) / len(vals)
        print(f"  盈提示={k:>10}: 样本={len(vals):>5} 涨概率={pos/len(vals)*100:>5.1f}% 平均={avg:>7.2f}")

print("\n=== 8. 漏提示 vs 下日涨跌 ===")
groups8 = defaultdict(list)
for row in all_rows:
    if row in forward_chg_1:
        lh = ws.cell(row, 93).value
        if lh:
            groups8[str(lh)].append(forward_chg_1[row])

for k, vals in sorted(groups8.items()):
    if len(vals) >= 3:
        pos = sum(1 for v in vals if v > 0)
        avg = sum(vals) / len(vals)
        print(f"  漏提示={k:>15}: 样本={len(vals):>5} 涨概率={pos/len(vals)*100:>5.1f}% 平均={avg:>7.2f}")

print("\n=== 9. 日周联动 vs 下日涨跌 ===")
groups9 = defaultdict(list)
for row in all_rows:
    if row in forward_chg_1:
        lk = ws.cell(row, 89).value
        if lk:
            groups9[str(lk)].append(forward_chg_1[row])

for k, vals in sorted(groups9.items()):
    if len(vals) >= 3:
        pos = sum(1 for v in vals if v > 0)
        avg = sum(vals) / len(vals)
        print(f"  联动={k:>10}: 样本={len(vals):>5} 涨概率={pos/len(vals)*100:>5.1f}% 平均={avg:>7.2f}")

print("\n=== 10. 仓月 vs 下周涨跌 ===")
groups10 = defaultdict(list)
for row in all_rows:
    if row in forward_chg_5:
        cm = ws.cell(row, 230).value
        if cm:
            groups10[str(cm)].append(forward_chg_5[row])

for k, vals in sorted(groups10.items()):
    if len(vals) >= 3:
        pos = sum(1 for v in vals if v > 0)
        avg = sum(vals) / len(vals)
        print(f"  仓月={k:>5}: 样本={len(vals):>5} 涨概率={pos/len(vals)*100:>5.1f}% 平均={avg:>7.2f}")

print("\n=== 11. 当前涨跌 vs 下日涨跌（动量效应）===")
# Group current change into deciles
chg_buckets = defaultdict(list)
for row in all_rows:
    if row in forward_chg_1:
        chg = ws.cell(row, 5).value
        if chg is not None:
            bucket = round(chg / 1) * 1  # group by 1% intervals
            chg_buckets[bucket].append((chg, forward_chg_1[row]))

for k in sorted(chg_buckets.keys()):
    vals = chg_buckets[k]
    if len(vals) >= 5:
        avg_fwd = sum(v[1] for v in vals) / len(vals)
        pos = sum(1 for v in vals if v[1] > 0)
        print(f"  今涨≈{k:>4}%: 样本={len(vals):>5} 明涨概率={pos/len(vals)*100:>5.1f}% 明日平均={avg_fwd:>7.2f}")

print("\n=== 12. 全局统计 ===")
all_1d = [forward_chg_1[r] for r in all_rows if r in forward_chg_1]
all_5d = [forward_chg_5[r] for r in all_rows if r in forward_chg_5]
all_20d = [forward_chg_20[r] for r in all_rows if r in forward_chg_20]

print(f"  下1日: 样本={len(all_1d)} 涨概率={sum(1 for v in all_1d if v>0)/len(all_1d)*100:.1f}% 平均={sum(all_1d)/len(all_1d):.2f}")
print(f"  下5日: 样本={len(all_5d)} 涨概率={sum(1 for v in all_5d if v>0)/len(all_5d)*100:.1f}% 平均={sum(all_5d)/len(all_5d):.2f}")
print(f"  下20日: 样本={len(all_20d)} 涨概率={sum(1 for v in all_20d if v>0)/len(all_20d)*100:.1f}% 平均={sum(all_20d)/len(all_20d):.2f}")
