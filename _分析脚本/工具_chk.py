import openpyxl

f = "算展D.sz300304.xlsx"
wb = openpyxl.load_workbook(f, read_only=True, data_only=True)

# Check sheets
print("Sheets:", wb.sheetnames)
ws = wb["LD300304"]

# Read row 1 headers
headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

# Find key columns
keys = ["柱排", "大局", "冲高提示", "乾坤", "仓周", "HA偏", "阳连", "鼎", "ZA", "ZB", "ZC", "ZD", "ZE", "仓指示", "仓位比", "柱型", "护型", "日周联动"]
for idx, h in enumerate(headers):
    if h and any(k in str(h) for k in keys):
        col = openpyxl.utils.get_column_letter(idx+1)
        print(f"  Col {col:3s} ({idx+1:3d}): {h}")

# Also check what BT columns look like
print("\n--- BT系列列名 ---")
for idx, h in enumerate(headers):
    if h and "BT" in str(h).upper():
        col = openpyxl.utils.get_column_letter(idx+1)
        print(f"  Col {col:3s} ({idx+1:3d}): {h}")

# Data row count
max_row = ws.max_row
print(f"\nTotal data rows: {max_row - 1}")
wb.close()
