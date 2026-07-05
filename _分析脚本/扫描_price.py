#!/usr/bin/env python3
"""
算展D 下日冲高概率分析脚本

分析条件:
- 下日涨跌幅 >= 3%（冲高）
- 日类BTZE/ZA/ZB/ZC/ZD 正负信号
- 日层护型 甲/乙/己 vs 丙/丁/戊
- 日层柱排 升/跌/人
- BTCD正交/负交
- DJE~DJC区间 / DJC~DJB区间

使用方法:
    python _分析脚本/check_price.py

依赖:
    pip install openpyxl
"""

import os
import sys
import openpyxl
from collections import defaultdict

# ──────────────────────────────────────────────
# 配置
# ──────────────────────────────────────────────
DATA_DIR = r"d:\@VSwork\VS昭明计划VBA优化"
FILES = [
    "算展D.sz300304.xlsx",
    "算展D.sz000510.xlsx",
    "算展D.sh600155.xlsx",
]
SURGE_THRESHOLD = 3.0  # 下日涨 >= 3%


def detect_columns(ws):
    """扫描第一行头，返回列号字典"""
    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_info = {}
    for ci, val in enumerate(row1):
        if val is not None:
            vs = str(val).replace("\n", "").replace(" ", "")
            col_info[vs] = ci  # 0-indexed

    # 输出所有列名（调试用）
    print(f"[列检测] 共 {len(col_info)} 个非空列")

    # 定位关键列
    cols = {}

    # 涨跌幅: 列名="涨"
    for vs, ci in col_info.items():
        if vs == "涨" and ci < 10:
            cols["涨"] = ci
            break

    # 日层护型: 搜索"护型DXAB"（列90=日层护型）
    for vs, ci in col_info.items():
        if "护型DXAB" in vs and 85 <= ci <= 95:
            cols["日层护型"] = ci
            break

    # 日层柱排: 搜索纯"柱排"且列号在~90-100
    for vs, ci in col_info.items():
        if vs == "柱排" and 90 <= ci <= 100:
            cols["日层柱排"] = ci
            break

    # 日类BTZ系列: 搜索 ZA_x000D_...日（族日均起始列=268）
    za_col = None
    for vs, ci in col_info.items():
        if "ZA_x000D_" in vs:
            za_col = ci
            break

    if za_col is not None:
        base = za_col
        cols["BTZA"] = base       # DJA (268)
        cols["BTZB"] = base + 1   # DJB (269)
        cols["BTZC"] = base + 2   # DJC (270)
        cols["BTAB"] = base + 3   # 护段AB (271)
        cols["BTAC"] = base + 4   # 护段AC (272)
        cols["BTBC"] = base + 5   # 护段BC (273, 可能无头但有数据)
        cols["BTZD"] = base + 6   # DJD (274)
        cols["BTZE"] = base + 7   # DJE (275)
        cols["BTCD"] = base + 8   # 护级CD (276)
        cols["BTCE"] = base + 9   # 护级CE (277)
        cols["BTDE"] = base + 10  # 护级DE (278)
        cols["族日均基"] = base

    # 总行数
    total_rows = 0
    for _ in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1

    return cols, total_rows


def safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def safe_int(v):
    if v is None:
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def get_zhu_prefix(zhu_val):
    """从柱排字符串提取前缀: 升/跌/人/错"""
    if zhu_val is None:
        return None
    zs = str(zhu_val)
    if zs.startswith("("):
        idx = zs.find(")")
        if idx > 0:
            return zs[1:idx]
    return zs[0]


def get_hu_type(hu_val):
    """从护型字符串提取第二字: 甲/乙/丙/丁/戊/己"""
    if hu_val is None:
        return None
    hs = str(hu_val)
    if len(hs) >= 2:
        return hs[1]
    return None


def analyze_one_file(filepath, stock_name):
    """分析单文件"""
    print(f"\n{'='*80}")
    print(f"分析文件: {stock_name}")
    print(f"文件路径: {filepath}")
    print(f"{'='*80}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    ld_sheet = None
    for sn in wb.sheetnames:
        if sn.startswith("LD"):
            ld_sheet = sn
            break
    if ld_sheet is None:
        print(f"  [错误] 未找到LD sheet")
        wb.close()
        return None, 0, 0, 0

    ws = wb[ld_sheet]
    print(f"  Sheet: {ld_sheet}")

    cols, total_rows = detect_columns(ws)
    print(f"\n  总行数: {total_rows}")
    print(f"  关键列:")
    for k in ["涨", "BTZA", "BTZB", "BTZC", "BTZE", "BTCD", "日层护型", "日层柱排"]:
        v = cols.get(k, None)
        print(f"    {k}: col={v}")

    required = ["涨", "BTZA", "BTZB", "BTZC", "BTZE", "BTCD"]
    missing = [r for r in required if r not in cols or cols[r] is None]
    if missing:
        print(f"  [错误] 缺少关键列: {missing}")
        wb.close()
        return None, 0, 0, 0

    # ── 读取数据 ──
    records = []
    row_num = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        if len(row) <= max(cols.values()):
            continue

        chang = safe_float(row[cols["涨"]])
        btza = safe_int(row[cols["BTZA"]])
        btzb = safe_int(row[cols["BTZB"]])
        btzc = safe_int(row[cols["BTZC"]])
        btze = safe_int(row[cols["BTZE"]])
        btcd = safe_int(row[cols["BTCD"]])

        hu_val = None
        if "日层护型" in cols and cols["日层护型"] is not None:
            hv = row[cols["日层护型"]]
            hu_val = str(hv) if hv is not None else None

        zhu_val = None
        if "日层柱排" in cols and cols["日层柱排"] is not None:
            zv = row[cols["日层柱排"]]
            zhu_val = str(zv) if zv is not None else None

        records.append({
            "row": row_num,
            "涨": chang,
            "BTZA": btza,
            "BTZB": btzb,
            "BTZC": btzc,
            "BTZE": btze,
            "BTCD": btcd,
            "护型": hu_val,
            "柱排": zhu_val,
            "柱排前缀": get_zhu_prefix(zhu_val),
            "护型第二字": get_hu_type(hu_val),
        })

    wb.close()
    print(f"  有效数据行: {len(records)}")

    # ── 计算下日收益 ──
    for i in range(len(records) - 1):
        records[i]["下日涨"] = records[i + 1]["涨"]
    records[-1]["下日涨"] = None

    valid = [r for r in records if r["下日涨"] is not None]
    total = len(valid)
    total_surge = sum(1 for r in valid if r["下日涨"] >= SURGE_THRESHOLD)
    total_expected = sum(r["下日涨"] for r in valid)

    print(f"\n  ── 基础统计 ──")
    print(f"  总样本数: {total}")
    print(f"  下日冲高(>={SURGE_THRESHOLD}%)次数: {total_surge}")
    print(f"  下日冲高概率: {total_surge/max(total,1)*100:.2f}%")
    print(f"  下日期望涨幅: {total_expected/max(total,1):.2f}%")

    # ── 条件概率 ──
    is_jia_yi_ji = lambda v: v is not None and v in ("甲", "乙", "己")
    is_bing_ding_wu = lambda v: v is not None and v in ("丙", "丁", "戊")

    conditions = [
        ("无条件(整体)", lambda r: True),
        ("BTZE>0(DJE正)", lambda r: r["BTZE"] is not None and r["BTZE"] > 0),
        ("BTZE>0 AND BTZA>0", lambda r: r["BTZE"] is not None and r["BTZE"] > 0 and r["BTZA"] is not None and r["BTZA"] > 0),
        ("BTZA>0(DJA正)", lambda r: r["BTZA"] is not None and r["BTZA"] > 0),
        ("BTZE>0 AND BTZA>0 AND 柱排=升", lambda r: r["BTZE"] is not None and r["BTZE"] > 0 and r["BTZA"] is not None and r["BTZA"] > 0 and r["柱排前缀"] == "升"),
        ("柱排=升", lambda r: r["柱排前缀"] == "升"),
        ("柱排=跌", lambda r: r["柱排前缀"] == "跌"),
        ("柱排=人", lambda r: r["柱排前缀"] == "人"),
        ("护型=甲/乙/己", lambda r: is_jia_yi_ji(r["护型第二字"])),
        ("护型=丙/丁/戊", lambda r: is_bing_ding_wu(r["护型第二字"])),
        ("护型=甲", lambda r: r["护型第二字"] == "甲"),
        ("护型=乙", lambda r: r["护型第二字"] == "乙"),
        ("护型=己", lambda r: r["护型第二字"] == "己"),
        ("护型=丙", lambda r: r["护型第二字"] == "丙"),
        ("护型=丁", lambda r: r["护型第二字"] == "丁"),
        ("BTCD>0(DXCD正交)", lambda r: r["BTCD"] is not None and r["BTCD"] > 0),
        ("BTCD<=0(DXCD负交)", lambda r: r["BTCD"] is not None and r["BTCD"] <= 0),
        ("DJE~DJC之间(BTZE<=0 AND BTZC>0)", lambda r: r["BTZE"] is not None and r["BTZE"] <= 0 and r["BTZC"] is not None and r["BTZC"] > 0),
        ("DJC~DJB之间(BTZC<=0 AND BTZB>0)", lambda r: r["BTZC"] is not None and r["BTZC"] <= 0 and r["BTZB"] is not None and r["BTZB"] > 0),
        ("BTZE<=0(DJE非正)", lambda r: r["BTZE"] is not None and r["BTZE"] <= 0),
        ("BTZA<=0(DJA非正)", lambda r: r["BTZA"] is not None and r["BTZA"] <= 0),
        ("BTZE>0 AND BTCD>0", lambda r: r["BTZE"] is not None and r["BTZE"] > 0 and r["BTCD"] is not None and r["BTCD"] > 0),
        ("柱排=升 AND 护型=甲", lambda r: r["柱排前缀"] == "升" and r["护型第二字"] == "甲"),
        ("BTZE>0 AND 柱排=升", lambda r: r["BTZE"] is not None and r["BTZE"] > 0 and r["柱排前缀"] == "升"),
        ("BTZE>0 AND 护型=甲/乙/己", lambda r: r["BTZE"] is not None and r["BTZE"] > 0 and is_jia_yi_ji(r["护型第二字"])),
        ("BTZE>0 AND 护型=甲", lambda r: r["BTZE"] is not None and r["BTZE"] > 0 and r["护型第二字"] == "甲"),
    ]

    print(f"\n  ── 下日冲高(>={SURGE_THRESHOLD}%)条件概率表 ──")
    print(f"  {'条件':<42s} {'样本数':>6s} {'冲高数':>6s} {'概率':>8s} {'期望涨幅':>8s}")
    print(f"  {'-'*70}")

    results = []
    for cond_name, cond_fn in conditions:
        subset = [r for r in valid if cond_fn(r)]
        n = len(subset)
        if n == 0:
            print(f"  {cond_name:<42s} {0:>6d} {'N/A':>6s} {'N/A':>8s} {'N/A':>8s}")
            results.append((cond_name, 0, 0, 0.0, 0.0))
            continue
        surge = sum(1 for r in subset if r["下日涨"] >= SURGE_THRESHOLD)
        prob = surge / n * 100
        exp_ret = sum(r["下日涨"] for r in subset) / n
        print(f"  {cond_name:<42s} {n:>6d} {surge:>6d} {prob:>7.2f}% {exp_ret:>7.2f}%")
        results.append((cond_name, n, surge, prob, exp_ret))

    print(f"  {'-'*70}")

    # ── 组合条件矩阵 ──
    print(f"\n  ── 组合条件矩阵 ──")

    ze_signs = [("BTZE>0", lambda r: r["BTZE"] is not None and r["BTZE"] > 0),
                ("BTZE<=0", lambda r: r["BTZE"] is not None and r["BTZE"] <= 0)]
    zhu_types = [("柱排升", lambda r: r["柱排前缀"] == "升"),
                 ("柱排跌", lambda r: r["柱排前缀"] == "跌"),
                 ("柱排人", lambda r: r["柱排前缀"] == "人"),
                 ("柱排无/错", lambda r: r["柱排前缀"] not in ("升", "跌", "人"))]
    hu_filter = [("护甲/乙/己", lambda r: is_jia_yi_ji(r["护型第二字"])),
                 ("护丙/丁/戊", lambda r: is_bing_ding_wu(r["护型第二字"])),
                 ("护无", lambda r: r["护型第二字"] is None)]

    # BTZE × 护型
    print(f"\n  [BTZE × 护型]")
    for z_name, z_fn in ze_signs:
        for h_name, h_fn in hu_filter:
            subset = [r for r in valid if z_fn(r) and h_fn(r)]
            n = len(subset)
            if n == 0:
                continue
            surge_n = sum(1 for r in subset if r["下日涨"] >= SURGE_THRESHOLD)
            prob_n = surge_n / n * 100
            exp_n = sum(r["下日涨"] for r in subset) / n
            print(f"    {z_name:<10s} + {h_name:<10s}: n={n:>5d}  surge={surge_n:>4d}  "
                  f"prob={prob_n:>6.2f}%  exp={exp_n:>6.2f}%")

    # BTZE × 柱排
    print(f"\n  [BTZE × 柱排]")
    for z_name, z_fn in ze_signs:
        for zh_name, zh_fn in zhu_types:
            subset = [r for r in valid if z_fn(r) and zh_fn(r)]
            n = len(subset)
            if n == 0:
                continue
            surge_n = sum(1 for r in subset if r["下日涨"] >= SURGE_THRESHOLD)
            prob_n = surge_n / n * 100
            exp_n = sum(r["下日涨"] for r in subset) / n
            print(f"    {z_name:<10s} + {zh_name:<10s}: n={n:>5d}  surge={surge_n:>4d}  "
                  f"prob={prob_n:>6.2f}%  exp={exp_n:>6.2f}%")

    # BTCD × 柱排
    print(f"\n  [BTCD × 柱排]")
    cd_signs = [("BTCD>0", lambda r: r["BTCD"] is not None and r["BTCD"] > 0),
                ("BTCD<=0", lambda r: r["BTCD"] is not None and r["BTCD"] <= 0)]
    for cd_name, cd_fn in cd_signs:
        for zh_name, zh_fn in zhu_types:
            subset = [r for r in valid if cd_fn(r) and zh_fn(r)]
            n = len(subset)
            if n == 0:
                continue
            surge_n = sum(1 for r in subset if r["下日涨"] >= SURGE_THRESHOLD)
            prob_n = surge_n / n * 100
            exp_n = sum(r["下日涨"] for r in subset) / n
            print(f"    {cd_name:<10s} + {zh_name:<10s}: n={n:>5d}  surge={surge_n:>4d}  "
                  f"prob={prob_n:>6.2f}%  exp={exp_n:>6.2f}%")

    # BTZA × 护型
    print(f"\n  [BTZA × 护型]")
    za_signs = [("BTZA>0", lambda r: r["BTZA"] is not None and r["BTZA"] > 0),
                ("BTZA<=0", lambda r: r["BTZA"] is not None and r["BTZA"] <= 0)]
    for za_name, za_fn in za_signs:
        for h_name, h_fn in hu_filter:
            subset = [r for r in valid if za_fn(r) and h_fn(r)]
            n = len(subset)
            if n == 0:
                continue
            surge_n = sum(1 for r in subset if r["下日涨"] >= SURGE_THRESHOLD)
            prob_n = surge_n / n * 100
            exp_n = sum(r["下日涨"] for r in subset) / n
            print(f"    {za_name:<10s} + {h_name:<10s}: n={n:>5d}  surge={surge_n:>4d}  "
                  f"prob={prob_n:>6.2f}%  exp={exp_n:>6.2f}%")

    return results, total, total_surge, total_expected / max(total, 1)


def main():
    print("=" * 80)
    print("  算展D 下日冲高概率分析")
    print(f"  冲高阈值: 下日涨 >= {SURGE_THRESHOLD}%")
    print(f"  分析文件数量: {len(FILES)}")
    print("=" * 80)

    all_file_results = []
    for fname in FILES:
        fpath = os.path.join(DATA_DIR, fname)
        if not os.path.exists(fpath):
            print(f"\n  [跳过] 文件不存在: {fpath}")
            continue
        stock_name = fname.replace("算展D.", "").replace(".xlsx", "")
        result = analyze_one_file(fpath, stock_name)
        if result[0] is not None:
            all_file_results.append((stock_name, result))

    # ── 三股汇总 ──
    if len(all_file_results) >= 2:
        print(f"\n\n{'='*80}")
        print(f"  多股汇总对比")
        print(f"{'='*80}")
        first_results = all_file_results[0][1][0]
        for ri, (cond_name, _, _, _, _) in enumerate(first_results):
            if ri % 3 == 0:
                print(f"\n  {'条件':<42s}", end="")
                for fname, _ in all_file_results:
                    print(f" {fname:>18s}", end="")
                print()
            print(f"  {cond_name:<42s}", end="")
            for fname, (results, total_cnt, surge_cnt, exp_rtn) in all_file_results:
                if ri < len(results):
                    _, n, s, p, e = results[ri]
                    if n > 0:
                        print(f" {p:>6.2f}%({n:>4d})", end="  ")
                    else:
                        print(f" {'N/A':>16s}", end="  ")
                else:
                    print(f" {'N/A':>16s}", end="  ")
            print()

    print(f"\n\n  分析完成。")


if __name__ == "__main__":
    main()