"""检查价格列精确结构"""
from openpyxl import load_workbook

wb = load_workbook("算展D.sz300304.xlsx", read_only=True, data_only=True)
ws = wb["LD300304"]
headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

# 检查列28-35的列名
for c in range(28, 36):
    print(f"Col {c}: {repr(headers[c-1][:60])}")

# 检查是否有开盘收盘价列
for idx, h in enumerate(headers):
    if h and any(k in str(h) for k in ["结收", "结开", "价C", "收盘", "开盘"]):
        print(f"Col {idx+1}: {repr(h[:60])}")

# 第5列=涨, 列118=PR, 列125=PC
# 检查是否有成交价格序列——看第2行和第100行
print("\n--- 多行数据对比 ---")
for r in [2, 3, 100, 500, 1000, 2000]:
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    print(f"Row{r}: 涨={row[4]}, PR={row[117]}, PC={row[124] if len(row)>124 else None}")

# 验证：列5=涨 vs 列118=PR 是否一致
cnt = 0
for r in range(2, 100):
    row = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
    if abs(row[4] - row[117]) > 0.001:
        cnt += 1
        print(f"Row{r}: 涨={row[4]} PR={row[117]} 不一致")
print(f"\n前99行中不一致: {cnt}行")

wb.close()
