import openpyxl
from collections import defaultdict

files = [
    ('算展D.sh600155.xlsx', 'sh600155'),
    ('算展D.sz300304.xlsx', 'sz300304'),
    ('算展D.sz000510.xlsx', 'sz000510'),
]

all_results = []

for path, label in files:
    print(f"\n{'='*70}")
    print(f"【{label}】加载中...")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['LD' + label[2:]]  # e.g. LD600155
    rows = list(range(2, ws.max_row + 1))
    print(f"  总行数: {len(rows)}")

    # 1. 基础统计
    total_days = len(rows)
    up_days = sum(1 for r in rows if ws.cell(r, 5).value and ws.cell(r, 5).value > 0)
    down_days = sum(1 for r in rows if ws.cell(r, 5).value and ws.cell(r, 5).value < 0)
    avg_chg = sum(ws.cell(r, 5).value or 0 for r in rows) / total_days
    print(f"  涨:{up_days}({up_days/total_days*100:.1f}%) 跌:{down_days}({down_days/total_days*100:.1f}%) 平均:{avg_chg:.2f}%")

    # 2. Build forward targets
    fwd_1d_chg = {}
    fwd_1d_high = {}
    for i, r in enumerate(rows):
        if i+1 < len(rows):
            nr = rows[i+1]
            # Next day change (PR=col 118, but also col 5)
            nchg = ws.cell(nr, 5).value
            if nchg is not None:
                fwd_1d_chg[r] = nchg
            # Next day high amplitude (HR=col 116)
            nhr = ws.cell(nr, 116).value
            if nhr is not None:
                fwd_1d_high[r] = nhr

    print(f"  有下日数据: {len(fwd_1d_chg)}行")

    # 3. Test each oracle feature against next-day "冲高" (next day change > 2%)
    def test_feature(col, name, rows, fwd, min_samples=10):
        groups = defaultdict(list)
        for r in rows:
            if r in fwd:
                v = ws.cell(r, col).value
                if v is not None and str(v).strip():
                    groups[str(v)].append(fwd[r])

        results = []
        for k, vals in groups.items():
            if len(vals) >= min_samples:
                pos = sum(1 for v in vals if v > 2.0)  # "冲高" = next day > 2%
                neg = sum(1 for v in vals if v < -2.0)  # "大跌" = next day < -2%
                avg = sum(vals) / len(vals)
                win_rate = pos / len(vals)
                results.append((k, len(vals), win_rate, avg, pos, neg))

        if results:
            results.sort(key=lambda x: x[2], reverse=True)
            return results
        return []

    print(f"\n  --- 关键特征 vs 下日涨幅>2%（冲高概率） ---")

    # 特征列表: (列号, 名称, 前缀截断长度)
    features = [
        (225, '仓周', 0),
        (226, '仓日', 0),
        (230, '仓月', 2),
        (88, '大局首字', 1),
        (79, '护型首字', 1),
        (77, '三鳄首字', 1),
        (92, '盈提示', 0),
        (93, '漏提示', 4),
        (89, '日周联动', 0),
        (86, '猪操作下周', 0),
    ]

    for col, fname, prefix_len in features:
        results = test_feature(col, fname, rows, fwd_1d_chg)
        if results:
            # Aggregate by prefix if specified
            if prefix_len > 0:
                agg = defaultdict(lambda: [0, 0, 0.0, 0])
                for k, n, wr, avg, pos, neg in results:
                    key = k[:prefix_len]
                    agg[key][0] += n
                    agg[key][1] += pos
                    agg[key][2] += avg * n
                    agg[key][3] += n
                print(f"\n  {fname}:")
                for k in sorted(agg.keys()):
                    v = agg[k]
                    total_n = v[0]
                    total_pos = v[1]
                    avg_avg = v[2] / v[3] if v[3] > 0 else 0
                    pct = total_pos / total_n * 100
                    print(f"    {k}: {total_n}次 冲高概率{pct:.1f}% 平均{v[2]/v[3]:.2f}%" if v[3] > 0 else "")
                    # Show top 3 sub-items
                    sub_results = [(k2, n, wr, avg2) for k2, n, wr, avg2, _, _ in results if k2[:prefix_len] == k]
                    sub_results.sort(key=lambda x: x[2], reverse=True)
                    for k2, n, wr, avg2 in sub_results[:3]:
                        print(f"      ├ {k2}: {n}次 冲高{wr*100:.0f}% 平均{avg2:.1f}%")
            else:
                print(f"\n  {fname}:")
                for k, n, wr, avg, pos, neg in results[:8]:
                    print(f"    {k}: {n}次 冲高概率{wr*100:.1f}% 平均{avg:.2f}%")

    # 4. Special: 下日涨跌 vs PR/HR今日动量
    print(f"\n  --- 今日涨跌 vs 下日冲高 ---")
    for threshold in [-5, -3, 0, 3, 5, 7]:
        subset = [(r, fwd_1d_chg[r]) for r in rows if r in fwd_1d_chg and ws.cell(r, 5).value and ws.cell(r, 5).value >= threshold]
        if threshold == 0:
            subset = [(r, fwd_1d_chg[r]) for r in rows if r in fwd_1d_chg and ws.cell(r, 5).value and abs(ws.cell(r, 5).value) < 0.5]
        if len(subset) > 10:
            pos = sum(1 for _, v in subset if v > 2)
            avg = sum(v for _, v in subset) / len(subset)
            print(f"    今日≈{threshold}%: {len(subset)}次 下日冲高>{pos/len(subset)*100:.1f}% 下日平均{avg:.2f}%")

    # 5. 星期效应
    from collections import Counter
    wd_groups = defaultdict(list)
    for r in rows:
        if r in fwd_1d_chg:
            wd = ws.cell(r, 4).value
            if wd:
                wd_groups[wd].append(fwd_1d_chg[r])
    print(f"\n  --- 星期 vs 下日冲高 ---")
    for wd in range(1, 6):
        if wd in wd_groups:
            vals = wd_groups[wd]
            pos = sum(1 for v in vals if v > 2)
            avg = sum(vals) / len(vals)
            days = ['','周一','周二','周三','周四','周五']
            print(f"    {days[wd]}: {len(vals)}次 冲高概率{pos/len(vals)*100:.1f}% 平均{avg:.2f}%")

    wb.close()
    print()
