"""快速批量分析 — openpyxl read_only 模式 (比COM快10倍)"""
import openpyxl, os, glob, time
from collections import defaultdict

DATA_DIR = r'D:\@VSwork\VS昭明计划VBA优化\昭明算展'
files = sorted(glob.glob(os.path.join(DATA_DIR, '算展.*.xlsx')))
print(f'文件数: {len(files)}', flush=True)

stats = {'DSHA': defaultdict(list), '鼎': defaultdict(list),
         '管宽': defaultdict(list), '四域日': defaultdict(list), '柱排': defaultdict(list)}
t0 = time.time()
processed = 0

for fi, fn in enumerate(files):
    try:
        wb = openpyxl.load_workbook(fn, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]

        # 读列头（第一行）
        headers = {}
        for c in range(1, 320):
            v = ws.cell(1, c).value
            if v: headers[str(v).replace('_x000D_','').replace('\r','').replace('\n','').replace(' ','')] = c

        c_idx = lambda n: headers.get(n)
        c_DSHA, c_鼎, c_K, c_S, c_Z, c_HR = c_idx('脸日'), c_idx('鼎日买提示'), c_idx('宽顶JC管顶JC'), c_idx('四域日'), c_idx('柱排'), c_idx('HR')

        prev = {'d':None,'ding':None,'k':None,'s':None,'z':None}
        for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row: continue
            def v(col):
                return row[col-1] if col and col <= len(row) else None

            hr = v(c_HR)
            if hr is None:
                prev['d'], prev['ding'], prev['k'], prev['s'], prev['z'] = v(c_DSHA), v(c_鼎), v(c_K), v(c_S), v(c_Z)
                continue
            try: hr = float(hr)
            except:
                prev['d'], prev['ding'], prev['k'], prev['s'], prev['z'] = v(c_DSHA), v(c_鼎), v(c_K), v(c_S), v(c_Z)
                continue

            try:
                if prev['d'] is not None:
                    vv = float(prev['d'])
                    b = '0~3' if vv<=3 else '3~5' if vv<=5 else '5~8' if vv<=8 else '8~10' if vv<=10 else '>10'
                    stats['DSHA'][b].append(hr)
                if prev['ding'] is not None:
                    vv = int(prev['ding'])
                    b = '1~2' if vv<=2 else '3~4' if vv<=4 else '5~6' if vv<=6 else '7~8' if vv<=8 else '≥9'
                    stats['鼎'][b].append(hr)
                if prev['k'] is not None:
                    vv = abs(float(prev['k']))
                    b = '0~5' if vv<=5 else '5~10' if vv<=10 else '10~20' if vv<=20 else '20~40' if vv<=40 else '>40'
                    stats['管宽'][b].append(hr)
                if prev['s'] is not None:
                    s = str(prev['s'])[:2]
                    if s: stats['四域日'][s].append(hr)
                if prev['z'] is not None:
                    z = str(prev['z'])
                    if z.startswith('升'): stats['柱排']['升'].append(hr)
                    elif z.startswith('跌'): stats['柱排']['跌'].append(hr)
                    elif '人' in z: stats['柱排']['人'].append(hr)
            except: pass

            prev['d'], prev['ding'], prev['k'], prev['s'], prev['z'] = v(c_DSHA), v(c_鼎), v(c_K), v(c_S), v(c_Z)

        wb.close()
        processed += 1
    except Exception as e:
        print(f'  ⚠️ {os.path.basename(fn)}: {e}', flush=True)

    if (fi+1)%20 == 0:
        print(f'  [{fi+1}/{len(files)}] {time.time()-t0:.0f}s', flush=True)

def pct(hrs, t=3):
    if not hrs: return 0, 0
    n = len(hrs)
    return sum(1 for h in hrs if h>=t)/n*100, n

def prn(title, data):
    print(f'\n{title}')
    print(f'{"区间":<8s} {"样本":>8s} {"P≥3%":>8s} {"P≥5%":>8s} {"均HR":>8s}')
    for k in sorted(data.keys()):
        h = data[k]; p3, n = pct(h,3); p5, _ = pct(h,5)
        print(f'{k:<8s} {n:>8d} {p3:>7.1f}% {p5:>7.1f}% {sum(h)/n:>7.2f}%')

print(f'\n✅ {processed}/{len(files)}个文件, {time.time()-t0:.0f}s', flush=True)
print('='*50)
prn('DSHA→后续冲高', stats['DSHA'])
prn('鼎→后续冲高', stats['鼎'])
prn('管宽→后续冲高', stats['管宽'])
prn('四域日→后续冲高', stats['四域日'])
prn('柱排→后续冲高', stats['柱排'])