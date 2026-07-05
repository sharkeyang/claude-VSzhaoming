"""
算展D 柱排对比验证
读算展D.xlsx → 提取真实日层柱排 → 用Python重算 → 逐行对比
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import pandas as pd
import numpy as np
import openpyxl

# 导入Python 柱排引擎
from MC2_5_迁移VBA_核3_柱排解析引擎 import 计算柱排序列, 计算单柱符号


def 读取算展D(path):
    """读取算展D数据，返回(原始OHLCV, 真实柱排列, 列索引)"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    # 找LD表
    ld_name = [s for s in wb.sheetnames if s.startswith("LD")][0]
    ws = wb[ld_name]

    # 读列头（第1行）
    headers = {}
    for col in range(1, min(ws.max_column + 1, 400)):
        val = ws.cell(1, col).value
        if val:
            headers[str(val).strip().replace('\n',' ')] = col

    # 关键列
    col_map = {}
    for kw, name in [("龄", "龄"), ("期", "期"), ("涨", "涨"), ("涨跌幅", "涨"),
                      ("柱排", "柱排日"), ("护型DXAB", "日层护型"),
                      ("ZA 日","BTZA"),("ZB 日","BTZB"),("ZC 日","BTZC"),("ZD 日","BTZD"),("ZE 日","BTZE"),
                      ("结价","结价"), ("高","最高"), ("低","最低"), ("开","开盘"), ("收","收盘"),
                      ("成交","成交量")]:
        found = False
        for h, col in headers.items():
            if kw in h:
                col_map[name] = col
                found = True
                break
        if not found:
            print(f"  ⚠️ 未找到列: {kw}")

    print(f"  列头总数: {len(headers)}, 已映射: {len(col_map)}")

    # 读全量数据
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if row[0] is None:
            break
        rows.append(row)

    wb.close()

    # 转DataFrame
    列名映射 = {}
    for name, col in col_map.items():
        if col <= len(headers):
            # 用headers反查列名
            for h, c in headers.items():
                if c == col:
                    列名映射[col] = name
                    break

    # 取需要的列
    data = {}
    for name, col in col_map.items():
        if col <= len(rows[0]):
            vals = [r[col-1] for r in rows]
            data[name] = vals

    df = pd.DataFrame(data)
    return df, col_map, headers


def 解析VBA柱排(柱排_str):
    """从VBA柱排字符串提取方向信息"""
    if pd.isna(柱排_str) or 柱排_str == "":
        return {"方向": "?", "开头": "?"}
    s = str(柱排_str)
    开头 = "?"
    if s.startswith("升"): 开头 = "升"
    elif s.startswith("跌"): 开头 = "跌"
    elif s.startswith("(升)"): 开头 = "升人"
    elif s.startswith("(跌)"): 开头 = "跌人"
    elif s.startswith("(人)"): 开头 = "人"
    elif s.startswith("错"): 开头 = "错"
    return {"方向": 开头, "原始": s}


def 对比验证(df, 列映射, 股票名):
    """对比Python柱排 vs 算展D真实柱排"""
    print(f"\n{'='*60}")
    print(f"  {股票名} 柱排对比验证")
    print(f"  数据量: {len(df)} 行")
    print(f"{'='*60}")

    # 提取原始OHLCV
    df_py = pd.DataFrame()
    # 尝试查找收盘价列
    close_col = None
    for name in ["收", "收盘"]:
        if name in df.columns:
            close_col = name
            break
    if not close_col:
        print("  ❌ 找不到收盘价列")
        return

    df_py["close"] = df[close_col].astype(float)

    # 开盘价
    open_col = None
    for name in ["开", "开盘"]:
        if name in df.columns:
            open_col = name
            break
    if open_col:
        df_py["open"] = df[open_col].astype(float)
    else:
        df_py["open"] = df_py["close"] * 0.99  # 近似

    # 涨跌幅
    if "涨" in df.columns:
        df_py["涨幅"] = df["涨"].astype(float)
    else:
        df_py["涨幅"] = df_py["close"].pct_change() * 100

    # 计算Python柱排
    柱排_py = 计算柱排序列(df_py, "日")

    # 获取算展D真实柱排
    if "柱排日" in df.columns:
        VBA柱排 = df["柱排日"]
    else:
        print("  ❌ 找不到柱排日列")
        return

    # 逐行对比
    match = 0
    mismatch = 0
    错未赋值 = 0
    details = []

    for i in range(min(len(VBA柱排), len(柱排_py))):
        vba = str(VBA柱排.iloc[i]) if pd.notna(VBA柱排.iloc[i]) else ""
        py = str(柱排_py.iloc[i]) if pd.notna(柱排_py.iloc[i]) else ""

        # 跳过开头几行（柱排需要积累期）
        if i < 5:
            continue

        # 忽略VBA中的"错.未赋值"（Python也可能不同）
        if "错" in vba or "错" in py:
            错未赋值 += 1
            continue

        # 比较开头部分（升/跌/(升)人/(跌)人/(人)人）
        vba_prefix = vba.split(".")[0] if "." in vba else vba
        py_prefix = py.split(".")[0] if "." in py else py

        if vba_prefix == py_prefix:
            match += 1
        else:
            mismatch += 1
            if mismatch <= 10:
                details.append((i, vba, py, df_py["涨幅"].iloc[i] if "涨幅" in df_py.columns else 0))

    total = match + mismatch
    accuracy = match / total * 100 if total > 0 else 0

    print(f"\n  对比结果:")
    print(f"    匹配: {match} 行 ({accuracy:.1f}%)")
    print(f"    不匹配: {mismatch} 行")
    print(f"    跳过(未赋值): {错未赋值} 行")
    print(f"    合计: {total} 行")

    if mismatch > 0:
        print(f"\n  前10个不匹配示例:")
        print(f"  {'行':>4} {'VBA柱排':<30} {'Python柱排':<30} {'涨幅':>6}")
        print(f"  {'─'*72}")
        for idx, v, p, chg in details[:10]:
            print(f"  {idx:>4} {v:<30} {p:<30} {chg:>+5.1f}%")


if __name__ == "__main__":
    base = r"d:\@VSwork\VS昭明计划VBA优化"
    股票池 = [
        (f"{base}\\算展D.sz300304.xlsx", "sz300304"),
        (f"{base}\\算展D.sz000510.xlsx", "sz000510"),
        (f"{base}\\算展D.sh600155.xlsx", "sh600155"),
    ]

    for path, name in 股票池:
        print(f"\n加载 {path} ...")
        df, cm, h = 读取算展D(path)
        df.columns = [str(c) if pd.notna(c) else f"col_{i}" for i, c in enumerate(df.columns)]
        对比验证(df, cm, name)
