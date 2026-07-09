"""
M2.1 data_loader.py — 从算展D LD Sheet 提取信号列

依据：M1.1_策略_指标映射表.md + 柱排量化规则.md
"""

import datetime
from pathlib import Path
import pandas as pd
import numpy as np

# ============================================================
# VBA常量 → Excel列号（基于算展D结构侦察）
# LD sheet: Row1=列名, Row2+ = 数据
# 列号 = openpyxl列索引(从1开始)
# ============================================================

class LDColumns:
    """算展D LD Sheet 关键列位置（索引从1开始）"""
    B日期 = 2           # 期（Excel序列号）
    E涨幅 = 5           # 涨跌幅

    # --- 周级信号 ---
    C柱排周 = 82        # "柱排 周" (CD)
    C大局WXCD = 88      # "大局(WXCD+WJC+WXAB) 周" (CJ)
    C冲高提示 = 87      # "下柱冲高提示 下周" (CI)
    C护型WXAB = 79      # "护型WXAB 周" (CA)
    C柱型WJA = 81       # "柱型WJA 周" (CC)

    # --- 日级信号 ---
    C柱排日 = 96        # "柱排" (CR)
    C柱型DJA = 94       # "柱型DJA" (CP)
    C日周联动 = 89      # "日周联动" (CK)
    C护型DXAB = 90      # "护型DXAB" (CL)

    # --- 日级BTZ系列（均线位次：正=线上，负=线下）---
    CZA日 = 268    # ZA 日 (JH) — DJA站稳
    CZB日 = 269    # ZB 日 (JI) — DJB
    CZC日 = 270    # ZC 日 (JJ) — DJC
    CZD日 = 274    # ZD 日 (JN) — DJD
    CZE日 = 273    # ZE 日 (JM) — DJE

    # --- 四域体系（列尾新增，列名匹配）---
    C四域周 = "四域周"
    C四域日 = "四域日"
    C日层段 = "日层段"
    C日机警 = "日机警"
    # --- 策策略列 ---
    C策长基仓 = "策长基仓"
    C策周浮仓 = "策周浮仓"
    C策日浮仓 = "策日浮仓"
    C策传 = "策传"

    # --- 周级BTZ（用于分清级别）---
    CZA周 = 282    # ZA 周 (JV)
    CZB周 = 283    # ZB 周 (JW)
    CZC周 = 284    # ZC 周 (JX)
    CZD周 = 288    # ZD 周 (KB)
    CZE周 = 292    # ZE 周 (KF)

    # --- 仓位信号 ---
    C乾坤路径 = 234     # "乾坤路径" (HZ)
    C仓指示 = 210       # "仓指示" (HB)
    C仓位比 = 212       # "仓位比" (HD)
    C仓周 = 225         # "仓周" (HQ)

    # --- 辅助 ---
    CHA偏 = 109         # "HA偏 日" (DE) — DSHA
    C鼎日 = 113         # "鼎 日买提示" (DI)
    C阳连日 = 115       # "阳连 日" (DK)
    C结价周 = 36        # "结价 周前" — 用于计算初始收盘价

    # --- 策分卡输入列 ---
    C盈提示周 = 85       # "盈提示 周"
    C波型周 = 80        # "波型（猪震）周"


# ============================================================
# 柱排字符串解析
# ============================================================

def parse_柱排(柱排_str: str) -> dict:
    """解析柱排字符串，返回结构化信息

    柱排格式: {方向}.{尾部形态}{符号序列}.{最后并符}{连数}
    例: "升.尾连QQ.Q3"  "跌.尾吞oV"  "(升)人.连后吞Qv"
    """
    result = {
        "方向": "",         # 升 / 跌 / 人
        "确定性": "确定",     # 确定 / 人工 / 无方向
        "尾部形态": "",      # 尾连 / 尾吞 / 尾反孕 / 连后吞 / 吞吞 / 孕孕
        "最后并符": "",      # Q(升连) / W(跌连) / O(升吞) / V(跌吞) / o(升孕) / v(跌孕)
        "连数": 0,          # 最后连续次数
        "是否是升开头": False,
        "是否是跌开头": False,
        "是人排": False,
    }

    if not 柱排_str or 柱排_str == "错.未赋值":
        result["方向"] = "未知"
        return result

    # 提取方向
    if 柱排_str.startswith("升"):
        result["方向"] = "升"
        result["是否是升开头"] = True
    elif 柱排_str.startswith("跌"):
        result["方向"] = "跌"
        result["是否是跌开头"] = True
    elif "(升)人" in 柱排_str:
        result["方向"] = "升"
        result["确定性"] = "人工"
        result["是人排"] = True
    elif "(跌)人" in 柱排_str:
        result["方向"] = "跌"
        result["确定性"] = "人工"
        result["是人排"] = True
    elif "(人)人" in 柱排_str or "人." in 柱排_str:
        result["方向"] = "无方向"
        result["确定性"] = "无方向"
        result["是人排"] = True
    else:
        result["方向"] = "未知"

    # 提取尾部形态
    for 形态 in ["尾连", "尾吞", "尾反孕", "连后吞", "吞吞", "孕孕"]:
        if 形态 in 柱排_str:
            result["尾部形态"] = 形态
            break

    # 提取最后并符和连数
    if ".Q" in 柱排_str:
        try:
            result["最后并符"] = "Q"
            连数_str = 柱排_str.split(".Q")[-1].split(".")[0]
            result["连数"] = abs(int(连数_str))
        except:
            pass
    elif ".W" in 柱排_str:
        try:
            result["最后并符"] = "W"
            连数_str = 柱排_str.split(".W")[-1].split(".")[0]
            result["连数"] = abs(int(连数_str))
        except:
            pass

    return result


def is_升链(柱排_str: str) -> bool:
    """判断柱排是否为可靠的升链"""
    p = parse_柱排(柱排_str)
    return p["是否是升开头"] and p["确定性"] == "确定" and p["尾部形态"] not in ("尾反孕",)


def is_跌链(柱排_str: str) -> bool:
    """判断柱排是否为可靠的跌链"""
    p = parse_柱排(柱排_str)
    return p["是否是跌开头"] and p["确定性"] == "确定"


def is_可忽略阴柱(柱排_str: str) -> bool:
    """镜面对称❸：判断DJA之上的阴柱是否可以忽略"""
    if not 柱排_str or 柱排_str == "错.未赋值":
        return True
    p = parse_柱排(柱排_str)
    if not p["是否是跌开头"]:
        return True          # 不是跌开头就忽略
    if p["确定性"] != "确定":
        return True          # 人工判断的跌开头也忽略（可能转）
    if p["尾部形态"] == "尾反孕":
        return True          # 反孕可能转升
    return False             # 确认为跌链，不可忽略


# ============================================================
# WXCD 字符串解析
# ============================================================

def parse_WXCD(大局_str: str) -> dict:
    """解析大局WXCD字符串，返回结构化信息

    示例值: "金升a↗.4蛀.Aa龙猪.初"  "屎降"  "银平"
    """
    result = {
        "品质": "",          # 金 / 银 / 唏 / 屎 / 尿
        "方向": "",          # 升 / 降 / 平
    }
    if not 大局_str:
        return result
    if "金" in 大局_str:
        result["品质"] = "金"
    elif "银" in 大局_str:
        result["品质"] = "银"
    elif "唏" in 大局_str:
        result["品质"] = "唏嘘"
    elif "屎" in 大局_str:
        result["品质"] = "屎"
    elif "尿" in 大局_str:
        result["品质"] = "尿"

    if "升" in 大局_str:
        result["方向"] = "升"
    elif "降" in 大局_str:
        result["方向"] = "降"
    elif "平" in 大局_str:
        result["方向"] = "平"

    return result


def is_WXCD_合格(WXCD: str) -> bool:
    """硬条件①：WXCD=金或银"""
    p = parse_WXCD(WXCD)
    return p["品质"] in ("金", "银")


# ============================================================
# Excel日期转换
# ============================================================

def excel_serial_to_date(val):
    """Excel序列号或datetime → datetime.date"""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (datetime.datetime, pd.Timestamp)):
        return val.date() if hasattr(val, "date") else val
    base = datetime.datetime(1899, 12, 30)
    return (base + datetime.timedelta(days=int(val))).date()


# ============================================================
# 主加载函数
# ============================================================

def load_ld_data(excel_path: str, sheet_name: str = None) -> pd.DataFrame:
    """从算展D LD Sheet加载完整数据

    Args:
        excel_path: 算展D文件路径
        sheet_name: LD sheet名（None=自动检测）

    Returns:
        DataFrame，包含所有策略列
    """
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=True, data_only=True)

    # 自动检测LD sheet
    if sheet_name is None:
        candidates = [s for s in wb.sheetnames if s.startswith("LD")]
        if not candidates:
            raise ValueError(f"找不到LD sheet in {excel_path}")
        sheet_name = candidates[0]

    ws = wb[sheet_name]
    print(f"加载 {excel_path} -> {sheet_name} ({ws.max_row - 1} 行数据)")

    # 列映射：列号(1-based) → 列名
    COL = LDColumns()
    cols_to_read = {
        "日期": COL.B日期,
        "涨幅": COL.E涨幅,
        "柱排周": COL.C柱排周,
        "柱排日": COL.C柱排日,
        "WXCD": COL.C大局WXCD,
        "冲高提示": COL.C冲高提示,
        "ZA日": COL.CZA日, "ZB日": COL.CZB日, "ZC日": COL.CZC日,
        "ZD日": COL.CZD日, "ZE日": COL.CZE日,
        "ZA周": COL.CZA周, "ZB周": COL.CZB周, "ZC周": COL.CZC周,
        "ZD周": COL.CZD周, "ZE周": COL.CZE周,
        "乾坤路径": COL.C乾坤路径,
        "HA偏": COL.CHA偏,
        "仓指示": COL.C仓指示,
        "仓位比": COL.C仓位比,
        "仓周": COL.C仓周,
        "鼎日": COL.C鼎日,
        "阳连日": COL.C阳连日,
        "柱型DJA": COL.C柱型DJA,
        "护型DXAB": COL.C护型DXAB,
        "日周联动": COL.C日周联动,
        "结价_周前": COL.C结价周,
        # 策分卡输入列
        "护型WXAB": COL.C护型WXAB,
        "盈提示周": COL.C盈提示周,
        "波型周": COL.C波型周,
    }

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # 空行停止
            break
        record = {}
        for col_name, col_idx in cols_to_read.items():
            record[col_name] = row[col_idx - 1] if len(row) >= col_idx else None
        data.append(record)

    wb.close()

    df = pd.DataFrame(data)

    # 日期转换
    df["日期"] = df["日期"].apply(excel_serial_to_date)

    # 数值转换
    for col in ["ZA日", "ZB日", "ZC日", "ZD日", "ZE日",
                "ZA周", "ZB周", "ZC周", "ZD周", "ZE周",
                "HA偏", "涨幅", "仓位比"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # 派生：收盘价序列（从涨幅反推，初始价用结价）
    初始价 = None
    for idx in range(0, min(50, len(data))):
        if data[idx].get("结价_周前") is not None:
            初始价 = float(data[idx]["结价_周前"])
            break
    当前价 = 初始价 or 10.0
    收盘价列表 = []
    for idx in range(len(data)):
        涨 = data[idx].get("涨幅", 0) or 0
        当前价 = 当前价 * (1 + 涨 / 100)
        收盘价列表.append(当前价)
    df["收盘价"] = 收盘价列表

    # 派生：硬条件指标
    df["WXCD合格"] = df["WXCD"].apply(lambda x: is_WXCD_合格(str(x)) if pd.notna(x) else False)
    df["前周WXZC>0"] = df["ZC周"] > 0
    df["今日DXZE>0"] = df["ZE日"] > 0
    df["DXZD>0"] = df["ZD日"] > 0  # 日条件本质

    # 派生：日主战场
    df["DJED安心"] = (df["ZE日"] > 0) & (df["ZD日"] > 0)

    # 派生：分清级别
    df["周级预警"] = df["ZA周"] < 0  # 周级走弱(WXZA<0)
    df["周级清仓"] = df["ZC周"] < 0  # 周级止损(WXZC<0)

    # 派生：日线视角——空头排列判定
    # DJEDC 空头排列 = DJE<DJD<DJC = ZE<ZD<ZC（BTZ值越小表示越在下方）
    # 简单判定：DJE之下即空头格局
    df["空头格局"] = df["ZE日"] < 0
    # 更精确的空头排列：ZE<ZD<ZC < 0（DJ<DE均空头）
    df["DJEDC空头排列"] = (df["ZE日"] < 0) & (df["ZD日"] < 0) & (df["ZC日"] < 0)
    # 空头止损触发（从严格到宽松）
    df["空头止损_DJE"] = df["ZE日"] > 0  # 上破DJE→趋势转多
    df["空头止损_DJD"] = df["ZD日"] > 0  # 上破DJD→放宽
    df["空头止损_DJC"] = df["ZC日"] > 0  # 上破DJC→起码

    # 派生柱排解析
    for col_name in ["柱排周", "柱排日"]:
        parsed = df[col_name].apply(lambda x: parse_柱排(str(x) if pd.notna(x) else ""))
        df[f"{col_name}_升开头"] = parsed.apply(lambda p: p["是否是升开头"])
        df[f"{col_name}_跌开头"] = parsed.apply(lambda p: p["是否是跌开头"])
        df[f"{col_name}_是人排"] = parsed.apply(lambda p: p["是人排"])
        df[f"{col_name}_尾部"] = parsed.apply(lambda p: p["尾部形态"])

    df["冲高有信号"] = df["冲高提示"].apply(lambda x: bool(x) if pd.notna(x) else False)

    return df


if __name__ == "__main__":
    df = load_ld_data("算展D.sz300304.xlsx")
    print(f"\n数据范围: {df['日期'].min()} ~ {df['日期'].max()}")
    print(f"总行数: {len(df)}")

    # 验证几个指标
    print(f"\n硬条件通过率:")
    print(f"  WXCD合格: {df['WXCD合格'].sum()}/{len(df)}")
    print(f"  前周WXZC>0: {df['前周WXZC>0'].sum()}/{len(df)}")
    print(f"  今日DXZE>0: {df['今日DXZE>0'].sum()}/{len(df)}")
    print(f"  三条件全过: {(df['WXCD合格'] & df['前周WXZC>0'] & df['今日DXZE>0']).sum()}/{len(df)}")

    print(f"\n柱排日样例行:")
    for i in range(5):
        print(f"  {df.iloc[i]['日期']}: 柱排日={df.iloc[i]['柱排日']}  升开头={df.iloc[i]['柱排日_升开头']}")

    print(f"\n柱排周样例行:")
    for i in range(5):
        print(f"  {df.iloc[i]['日期']}: 柱排周={df.iloc[i]['柱排周']}  升开头={df.iloc[i]['柱排周_升开头']}")
