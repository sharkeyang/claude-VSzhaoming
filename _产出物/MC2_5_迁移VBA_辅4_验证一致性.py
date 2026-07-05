"""
迁移VBA程序数据一致性验证
对指定算展D文件，逐行对比：Python计算结果 vs Excel列原始值
"""
import sys, os, random
sys.path.insert(0, os.path.dirname(__file__))
import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path
from MC2_5_迁移VBA_核1_数据加载引擎 import 计算均线, 计算全部BTZ, 计算涨跌幅
from MC2_5_迁移VBA_核2_策略计算引擎 import 提取WXCD前缀, DXAB护型分类, 计算日层联动系列
from MC2_5_迁移VBA_核3_柱排解析引擎 import 计算柱排序列

BASE = r"d:\@VSwork\VS昭明计划VBA优化"
算展目录 = Path(BASE) / "昭明算展"
random.seed(42)
所有文件 = sorted(算展目录.glob("算展D.*.xlsx"))
选中文件 = random.sample(所有文件, 10)
print(f"共 {len(所有文件)} 个算展文件，随机选中10个:\n")
for f in 选中文件:
    print(f"  {f.name}")
print()

def 清洗列名(原始):
    s = str(原始).replace('\n', ' ').replace('\r', ' ').replace('_x000D_', ' ')
    while '  ' in s: s = s.replace('  ', ' ')
    return s.strip()

def 读取算展D(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    表名 = [s for s in wb.sheetnames if s.startswith("LD")][0]
    ws = wb[表名]
    # 列头 {列号: 清洗后列名}
    列头 = {col: 清洗列名(ws.cell(1, col).value) for col in range(1, ws.max_column+1) if ws.cell(1, col).value}
    # 数据
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None: break
        rows.append(row)
    wb.close()
    n_cols = len(rows[0]) if rows else 0

    def 取列号(关键字):
        found = None
        for ci, cn in 列头.items():
            if 关键字 in cn: found = ci
        return found

    # OHLCV
    ohlcv = pd.DataFrame()
    for kw, col in [("龄","age"),("期","period"),("涨跌幅","涨跌幅"),("开","open"),("收","close"),("高","high"),("低","low"),("成交","volume")]:
        ci = 取列号(kw)
        if ci and ci <= n_cols: ohlcv[col] = pd.to_numeric([r[ci-1] for r in rows], errors='coerce')
        else: ohlcv[col] = np.nan

    # VBA结算列
    vba = pd.DataFrame()
    vba_map = {"ZE 日":"BTZE","ZD 日":"BTZD","ZC 日":"BTZC","ZB 日":"BTZB","ZA 日":"BTZA",
               "AB 日":"BTAB","AC 日":"BTAC","CD 日":"BTCD","CE 日":"BTCE","DE 日":"BTDE",
               "柱排":"柱排日","护型DXAB":"日层护型"}
    for kw, vn in vba_map.items():
        ci = 取列号(kw)
        if ci and ci <= n_cols: vba[vn+"_VBA"] = [r[ci-1] for r in rows]
        else: vba[vn+"_VBA"] = np.nan

    # 大局(WXCD前缀) — 从 "大局(WXCD+WJC+WXAB) 周" 列取首字
    ci = 取列号("大局")
    if ci and ci <= n_cols: vba["大局_VBA"] = [r[ci-1] for r in rows]
    else: vba["大局_VBA"] = np.nan

    return ohlcv, vba


def 对比(py_vals, vba_vals, skip_head=5):
    n = min(len(py_vals), len(vba_vals))
    m = mm = s = 0
    det = []
    for i in range(n):
        if i < skip_head: s += 1; continue
        pv = py_vals.iloc[i] if hasattr(py_vals, 'iloc') else py_vals[i]
        vv = vba_vals.iloc[i] if hasattr(vba_vals, 'iloc') else vba_vals[i]
        if pd.isna(pv) or str(pv).strip()=='': s += 1; continue
        if pd.isna(vv) or str(vv).strip()=='': s += 1; continue
        vs = str(vv).strip()
        if "错" in vs: s += 1; continue  # VBA未赋值
        ps = str(pv).strip()
        if ps == vs: m += 1
        else:
            mm += 1
            if len(det) < 5: det.append((i, vs[:45], ps[:45]))
    total = m + mm
    acc = m/total*100 if total>0 else 0
    return m, mm, s, acc, det


def 验证单个文件(path):
    股票名 = path.stem.replace("算展D.", "")
    print(f"\n{'='*60}\n  {股票名}\n{'='*60}")
    ohlcv, vba = 读取算展D(path)
    n = len(ohlcv)
    print(f"  数据行: {n}")
    if n < 10: print("  SKIP"); return

    # --- Python计算 ---
    df = ohlcv[["close","open","high","low","volume"]].copy()
    df = 计算涨跌幅(df); df = 计算均线(df); df = 计算全部BTZ(df)
    df["WXCD"] = 提取WXCD前缀(df)
    df["护型"] = DXAB护型分类(df)
    df["联动"] = df["WXCD"] + 计算日层联动系列(
        df["BTZE"].values, df["BTCD"].values, df["BTZC"].values,
        df["BTBC"].values, df["BTAB"].values, df["BTZA"].values,
        df["BTZB"].values, df["BTZD"].values)
    df["柱排"] = 计算柱排序列(df, "日")

    # --- BTZ对比 ---
    for col in ["BTZE","BTZD","BTZC","BTZB","BTZA","BTAB","BTAC","BTCD","BTCE","BTDE"]:
        vk = col+"_VBA"
        if vk not in vba.columns or vba[vk].isna().all(): continue
        try: vba_int = vba[vk].apply(lambda x: int(float(x)) if pd.notna(x) else np.nan)
        except: continue
        _m,_mm,_s,a,_d = 对比(df[col].astype(int), vba_int)
        sym = "⚠️" if a < 97 else "✅"
        print(f"    {sym} {col}: {a:.1f}% ({_m}/{_m+_mm})")

    # --- WXCD对比 (VBA首字 = 金/银/唏/嘘/屎/尿) ---
    if "大局_VBA" in vba.columns:
        vba_wx = vba["大局_VBA"].apply(lambda x: str(x)[0] if pd.notna(x) else "")
        _m,_mm,_s,a,_d = 对比(df["WXCD"], vba_wx)
        sym = "⚠️" if a < 97 else "✅"
        print(f"    {sym} WXCD: {a:.1f}% ({_m}/{_m+_mm})")
        if _d:
            for idx,vb,py in _d[:3]:
                print(f"      row {idx}: VBA={vb}  PY={py}")

    # --- 护型对比 (VBA第二位 = 甲/乙/丙/丁/戊/己) ---
    if "日层护型_VBA" in vba.columns:
        vba_hx = vba["日层护型_VBA"].apply(lambda x: str(x)[1] if pd.notna(x) and len(str(x))>1 else "")
        _m,_mm,_s,a,_d = 对比(df["护型"], vba_hx)
        sym = "⚠️" if a < 80 else "✅"
        print(f"    {sym} 护型: {a:.1f}% ({_m}/{_m+_mm})")
        if _d:
            for idx,vb,py in _d[:3]:
                print(f"      row {idx}: VBA={vb}  PY={py}")

    # --- 柱排日对比 ---
    if "柱排日_VBA" in vba.columns:
        _m,_mm,_s,a,_d = 对比(df["柱排"], vba["柱排日_VBA"])
        sym = "⚠️" if a < 95 else "✅"
        print(f"    {sym} 柱排日: {a:.1f}% ({_m}/{_m+_mm})")
        if _d:
            for idx,vb,py in _d[:3]:
                print(f"      row {idx}: VBA={vb}  PY={py}")


for f in 选中文件:
    try: 验证单个文件(f)
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f"\n  ❌ {f.name}: {e}")

print(f"\n{'='*60}\n  验证完成 — 共检查 {len(选中文件)} 个文件")
