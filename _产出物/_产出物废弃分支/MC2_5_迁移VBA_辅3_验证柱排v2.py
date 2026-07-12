"""
算展D 柱排对比验证 v2
直接从算展D读取 并符串 → Python分类 → 对比VBA柱排
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from MC2_5_迁移VBA_核3_柱排解析引擎 import 柱排分类, 柱排后缀


def 遍历算展D(path):
    """读取算展D：并符串 + 真实柱排 + 涨跌幅"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ld_name = [s for s in wb.sheetnames if s.startswith("LD")][0]
    ws = wb[ld_name]

    # 列头定位
    headers = {}
    for col in range(1, 300):
        v = ws.cell(1, col).value
        if v:
            s = str(v).replace("\n"," ").replace("\r","").strip()
            if "并符串" in s and "日" in s: headers["并符串日"] = col
            if s == "柱排": headers["柱排日"] = col  # 精确匹配
            if s.startswith("涨"): headers["涨"] = col

    # 读取数据
    并符串日, 柱排日, 涨 = [], [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None: break
        if "并符串日" in headers:
            并符串日.append(str(row[headers["并符串日"]-1]) if row[headers["并符串日"]-1] else "")
        if "柱排日" in headers:
            柱排日.append(str(row[headers["柱排日"]-1]) if row[headers["柱排日"]-1] else "")

    wb.close()
    return 并符串日, 柱排日, headers


def 对比(并符串列表, 柱排列表, 股票名):
    """Python柱排分类 vs VBA柱排"""
    print(f"\n{'='*60}")
    print(f"  {股票名} 柱排对比")
    print(f"  数据: {len(并符串列表)} 行, 并符串列: {len(并符串列表[0]) if 并符串列表 else 0}")
    print(f"{'='*60}")

    match, mismatch, skip = 0, 0, 0
    细节 = []

    for i in range(len(并符串列表)):
        bf = 并符串列表[i]
        vba = 柱排列表[i] if i < len(柱排列表) else ""

        if not bf or bf == "None":
            skip += 1
            continue

        # Python分类
        py = 柱排分类(bf)

        # 不对比 "错.未赋值" 行
        if "错" in vba and "错" in py:
            skip += 1
            continue

        # 对比：忽略后缀（.W{N}/.Q{N}）只比前半部分
        vba_base = vba.split(".")[0] if "." in vba else vba
        py_base = py.split(".")[0] if "." in py else py
        # 忽略后缀差异，只比主要分类
        vba_core = vba_base
        py_core = py_base
        # 进一步简化: "(升)人.连后吞" → "(升)人"
        if "人" in vba_core and "人" in vba_core:
            vba_core = vba_core.split(".", 1)[0] if "." in vba_core else vba_core
            py_core = py_core.split(".", 1)[0] if "." in py_core else py_core

        if vba_base == py_base:
            match += 1
        elif vba_core == py_core:
            match += 1  # core分类一致，后缀尾缀略有差异也接受
        else:
            mismatch += 1
            if mismatch <= 8:
                细节.append((i, bf, vba, py))

    total = match + mismatch
    rate = match / total * 100 if total > 0 else 0

    print(f"  匹配: {match} ({rate:.1f}%)")
    print(f"  不匹配: {mismatch}")
    print(f"  跳过: {skip}")
    print(f"  有效: {total}")

    if 细节:
        print(f"\n  不匹配示例:")
        print(f"  {'行':>4} {'并符串':<12} {'VBA柱排':<30} {'Python':<30}")
        print(f"  {'-'*80}")
        for idx, bf, v, p in 细节:
            print(f"  {idx:>4} {bf:<12} {v:<30} {p:<30}")


if __name__ == "__main__":
    base = r"d:\@VSwork\VS昭明计划VBA优化"
    for f, name in [("算展D.sz300304.xlsx", "sz300304"),
                     ("算展D.sz000510.xlsx", "sz000510"),
                     ("算展D.sh600155.xlsx", "sh600155")]:
        bf, zp, h = 遍历算展D(f"{base}\\{f}")
        print(f"\n已定位列: {h}")
        对比(bf, zp, name)