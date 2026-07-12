"""
批处理验证：随机20只算展D文件，Python重算后逐项对比

流程:
  1. 从昭明算展/ 随机选20个文件
  2. 对每个文件:
     a. 从D:\zdata\data中股\{代码}.csv 读OHLCV
     b. 全量结算（EMA+BTZ+涨跌幅）
     c. 策略计算（WXCD+护型+联动+柱排）
     d. 从原始算展D.xlsx读取VBA计算列
     e. 逐列对比，统计一致率
  3. 汇总输出对比表
"""

import sys, os, random
sys.path.insert(0, os.path.dirname(__file__))

import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path
from datetime import datetime

from MC2_5_迁移VBA_核1_数据加载引擎 import 全量结算, 读取ZdataCSV
from MC2_5_迁移VBA_核2_策略计算引擎 import 提取WXCD前缀, DXAB护型分类, 计算日层联动系列
from MC2_5_迁移VBA_核3_柱排解析引擎 import 计算柱排序列

import warnings
warnings.filterwarnings("ignore")

random.seed(42)

BASE = r"D:\@VSwork\VS昭明计划VBA优化"
算展目录 = Path(BASE) / "昭明算展"
ZDATA目录 = r"D:\zdata\data中股"

# BTZ名映射
BTZ_VBA名 = {
    "BTZA": "ZA", "BTZB": "ZB", "BTZC": "ZC",
    "BTZD": "ZD", "BTZE": "ZE",
    "BTAB": "AB", "BTAC": "AC", "BTBC": "BC",
    "BTAD": "AD", "BTBD": "BD", "BTCD": "CD",
    "BTCE": "CE", "BTDE": "DE",
}

# 列扫描关键字 → 标准列名
VBA列扫描 = {
    "ZA 日": "ZA", "ZB 日": "ZB", "ZC 日": "ZC",
    "ZD 日": "ZD", "ZE 日": "ZE",
    "AB 日": "AB", "AC 日": "AC",
    "CD 日": "CD", "CE 日": "CE", "DE 日": "DE",
    "柱排": "柱排VBA",       # 日柱排
    "护型": "护型VBA",       # DXAB护型
    "大局": "WXCD_VBA",      # 大局(WXCD+WJC+WXAB) 周
}

# 日BTZ在VBA文件中的列号（基于M2_1_数据加载的映射）
BTZ日列号 = {"ZA": 268, "ZB": 269, "ZC": 270, "ZD": 274, "ZE": 275,
             "AB": 271, "AC": 272, "CD": 276, "CE": 277, "DE": 278}

# VBA策略列号
VBA策略列 = {"柱排": 96, "护型": 90, "大局": 88, "涨": 5}


def 清洗列名(原始):
    s = str(原始).replace('\n', ' ').replace('\r', ' ').replace('_x000D_', ' ')
    while '  ' in s: s = s.replace('  ', ' ')
    return s.strip()


def 读取VBA列(path):
    """从VBA算展D.xlsx读取关键列"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    表名 = [s for s in wb.sheetnames if s.startswith("LD")][0]
    ws = wb[表名]

    # 列头索引
    列头 = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            列头[c] = 清洗列名(v)

    # 读数据
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None: break
        rows.append(row)
    wb.close()
    n_cols = len(rows[0]) if rows else 0

    def 取列(关键字, 默认=None):
        for ci, cn in 列头.items():
            if 关键字 in cn and ci <= n_cols:
                return [r[ci-1] for r in rows]
        return 默认

    # 提取
    result = {}
    result["涨"] = 取列("涨")
    for kw, name in VBA列扫描.items():
        vals = 取列(kw)
        if vals:
            result[name] = vals

    # BTZ特定列号
    for btz_name, col_n in BTZ日列号.items():
        if col_n <= n_cols:
            result["BTZ_" + btz_name] = [r[col_n-1] for r in rows]

    return result, len(rows), 列头


def 对比列(py_vals, vba_vals, skip_head=3):
    """逐行对比Python vs VBA值，返回(匹配数, 总有效数, 一致率)"""
    n = min(len(py_vals), len(vba_vals))
    match = total = 0
    mismatches = []
    for i in range(skip_head, n):
        pv = py_vals.iloc[i] if hasattr(py_vals, 'iloc') else py_vals[i]
        vv = vba_vals.iloc[i] if hasattr(vba_vals, 'iloc') else vba_vals[i]
        # 跳过空值
        if pd.isna(pv) or str(pv).strip() in ('', '错.未赋值'):
            continue
        if pd.isna(vv) or str(vv).strip() in ('', '错.未赋值'):
            continue
        ps = str(pv).strip()
        vs = str(vv).strip()
        # 数值类: 转为int比较
        try:
            pi = int(float(pv))
            vi = int(float(vv))
            if pi == vi: match += 1
            else:
                if len(mismatches) < 3:
                    mismatches.append((i, vi, pi))
            total += 1
        except (ValueError, TypeError):
            # 字符串类: 直接比较
            if ps == vs: match += 1
            else:
                if len(mismatches) < 3:
                    mismatches.append((i, vs[:25], ps[:25]))
            total += 1
    acc = match / total * 100 if total > 0 else 0
    return match, total, acc, mismatches


def 验证单个(代码: str, vba_path: Path):
    """验证一只股票的Python vs VBA一致性"""
    print(f"\n{'─'*55}")
    print(f"  {代码}")
    print(f"{'─'*55}")

    # 1. 从zdata读原始数据
    try:
        df = 读取ZdataCSV(代码, ZDATA目录)
    except Exception as e:
        print(f"  ❌ zdata读取失败: {e}")
        return None
    print(f"  zdata: {len(df)}行 ({df['date'].min().date()}~{df['date'].max().date()})")

    # 2. 全量结算
    df = 全量结算(df)
    # 重命名BTZ
    for py_col, vba_col in BTZ_VBA名.items():
        if py_col in df.columns:
            df[vba_col] = df[py_col].astype(int)

    # 3. 策略计算
    df["WXCD"] = 提取WXCD前缀(df)
    df["护型"] = DXAB护型分类(df)
    df["联动"] = df["WXCD"] + 计算日层联动系列(
        df["ZE"].values, df["CD"].values, df["ZC"].values,
        df["BC"].values, df["AB"].values, df["ZA"].values,
        df["ZB"].values, df["ZD"].values)
    df["柱排"] = 计算柱排序列(df, "日")

    # 4. 读取VBA数据
    vba_data, vba_rows, vba_headers = 读取VBA列(vba_path)
    print(f"  VBA:   {vba_rows}行")

    # 5. 逐列对比
    results = {}

    # BTZ对比
    for btz_name in ["ZA", "ZB", "ZC", "ZD", "ZE", "AB", "AC", "CD", "CE", "DE"]:
        py_key = btz_name
        vba_key = "BTZ_" + btz_name
        if py_key not in df.columns or vba_key not in vba_data:
            continue
        m, t, acc, det = 对比列(df[py_key].astype(int), vba_data[vba_key])
        results[f"BTZ{btz_name}"] = (m, t, acc)

    # WXCD对比（VBA大局首字）
    if "WXCD_VBA" in vba_data:
        vba_wxcd = pd.Series([str(x)[0] if pd.notna(x) and len(str(x)) > 0 else "" for x in vba_data["WXCD_VBA"]])
        m, t, acc, det = 对比列(df["WXCD"], vba_wxcd)
        results["WXCD"] = (m, t, acc)

    # 护型对比
    if "护型VBA" in vba_data:
        vba_hx = pd.Series([str(x)[1] if pd.notna(x) and len(str(x)) > 1 else "" for x in vba_data["护型VBA"]])
        m, t, acc, det = 对比列(df["护型"], vba_hx)
        results["护型"] = (m, t, acc)

    # 柱排对比
    if "柱排VBA" in vba_data:
        m, t, acc, det = 对比列(df["柱排"], vba_data["柱排VBA"])
        results["柱排"] = (m, t, acc)

    # 涨跌幅对比
    if "涨" in vba_data and "涨幅" in df.columns:
        vba_涨 = pd.Series(vba_data["涨"])
        m, t, acc, det = 对比列(df["涨幅"], vba_涨)
        results["涨跌幅"] = (m, t, acc)

    return results


def main():
    # 随机选20个VBA算展文件
    所有文件 = sorted(算展目录.glob("算展D.*.xlsx"))
    random.seed(42)
    选中的 = random.sample(所有文件, 20)

    print("=" * 55)
    print(f"  VBA vs Python 一致性验证")
    print(f"  随机选中20个算展文件")
    print("=" * 55)
    for f in 选中的:
        print(f"  {f.name}")

    # 逐文件验证
    汇总 = []
    for fp in 选中的:
        代码 = fp.stem.replace("算展D.", "")
        results = 验证单个(代码, fp)
        if results:
            汇总.append((代码, results))
            # 打印该文件结果
            for col, (m, t, acc) in sorted(results.items()):
                sym = "✅" if acc >= 95 else ("⚠️" if acc >= 80 else "❌")
                print(f"  {sym} {col}: {acc:.1f}% ({m}/{t})")

    # 汇总表
    print(f"\n\n{'='*90}")
    print(f"  一致性验证汇总 — 20只股票")
    print(f"{'='*90}")

    # 所有指标列
    all_cols = ["涨跌幅", "BTZZA", "BTZZB", "BTZZC", "BTZZD", "BTZZE",
                "BTZAB", "BTZAC", "BTZCD", "BTZCE", "BTZDE",
                "WXCD", "护型", "柱排"]

    # 表头
    header = f"  {'代码':<10}"
    for col in all_cols:
        header += f" {col:>8}"
    print(header)
    print(f"  {'─'*10} {'─'*8}" * (1 + len(all_cols)))

    # 每行
    总分 = {col: [0, 0] for col in all_cols}
    for 代码, results in 汇总:
        line = f"  {代码:<10}"
        for col in all_cols:
            if col in results:
                m, t, acc = results[col]
                line += f" {acc:>7.1f}%"
                总分[col][0] += m
                总分[col][1] += t
            else:
                line += f" {'N/A':>8}"
        print(line)

    # 平均行
    avg_line = f"  {'平均':<10}"
    for col in all_cols:
        if 总分[col][1] > 0:
            avg = 总分[col][0] / 总分[col][1] * 100
            avg_line += f" {avg:>7.1f}%"
        else:
            avg_line += f" {'N/A':>8}"
    print(f"  {'─'*10} {'─'*8}" * (1 + len(all_cols)))
    print(avg_line)

    # 总体统计
    总匹配 = sum(v[0] for v in 总分.values())
    总有效 = sum(v[1] for v in 总分.values())
    print(f"\n  总有效对比: {总有效} 行")
    print(f"  总匹配数:   {总匹配} 行")
    print(f"  总一致率:   {总匹配/总有效*100:.1f}%")

    print(f"\n{'='*90}")


if __name__ == "__main__":
    main()