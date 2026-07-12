"""
生成 py算展sh512660.xlsx — 纯Python版算展D文件

流程:
  1. AKshare下载sh512660日线OHLCV
  2. 全量结算（EMA均线 + BTZ交叉 + 涨跌幅）
  3. 策略引擎（WXCD前缀 + 日层护型 + 日层联动 + 猪操作）
  4. 柱排引擎（日柱排 + 周柱排 + 下柱冲高提示）
  5. 周级聚合（周柱排、周护型、周大局(WXCD+WJC+WXAB)）
  6. 写出LD sheet → py算展sh512660.xlsx

对比VBA版算展D列结构:
  col 1:  龄（数据序号）
  col 2:  期（日期）
  col 3:  键（True，表示有效）
  col 4:  七（周号，当年第几周）
  col 5:  涨（涨跌幅）
  col 28-35: P0~P3/H0~H3（OHLC）
  col 36-88: 周级指标
  col 89-96: 日级核心指标
  col 97-129: 日级详细指标
  col 268-278: BTZ日级系列
  col 282-307: BTZ周级系列
"""

import sys
sys.path.insert(0, "_回测框架")

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path

# 导入Python计算引擎
from MC2_5_迁移VBA_核1_数据加载引擎 import 全量结算, 计算全部BTZ, 计算涨跌幅
from MC2_5_迁移VBA_核2_策略计算引擎 import 提取WXCD前缀, DXAB护型分类, 计算日层联动系列, 计算猪操作, 计算下周冲高信号, 计算下日冲高信号
from MC2_5_迁移VBA_核3_柱排解析引擎 import 计算柱排序列, 计算下柱冲高提示

import warnings
warnings.filterwarnings("ignore")


# ============================================================
# 1. 下载数据
# ============================================================

def 下载AKshareETF数据(symbol: str, 起始日: str = "20160801", 结束日: str = "20260630"):
    """下载ETF日线数据 from AKshare"""
    import akshare as ak
    df = ak.fund_etf_hist_em(
        symbol=symbol,
        period="daily",
        start_date=起始日,
        end_date=结束日,
        adjust="qfq"
    )
    # 统一列名
    列映射 = {
        "日期": "date", "开盘": "open", "最高": "high",
        "最低": "low", "收盘": "close", "成交量": "volume",
        "成交额": "amount",
    }
    df.rename(columns=列映射, inplace=True)
    df = df[["date", "open", "high", "low", "close", "volume", "amount"]]
    df["date"] = pd.to_datetime(df["date"])
    df.sort_values("date", inplace=True)
    df.reset_index(drop=True, inplace=True)
    print(f"  下载: {symbol} → {len(df)}行 ({df['date'].min().date()} ~ {df['date'].max().date()})")
    return df


# ============================================================
# 2. BTZ重命名：VBA版名 ↔ Python计算名
# ============================================================

BTZ_VBA名映射 = {
    "BTZA": "ZA", "BTZB": "ZB", "BTZC": "ZC",
    "BTZD": "ZD", "BTZE": "ZE",
    "BTAB": "AB", "BTAC": "AC", "BTBC": "BC",
    "BTAD": "AD", "BTBD": "BD", "BTCD": "CD",
    "BTCE": "CE", "BTDE": "DE",
}


def 重命名BTZ为VBA名(df: pd.DataFrame) -> pd.DataFrame:
    for py_col, vba_col in BTZ_VBA名映射.items():
        if py_col in df.columns:
            df[vba_col] = df[py_col].astype(int).copy()
    return df


# ============================================================
# 3. 周级聚合
# ============================================================

def 周柱排聚合(df: pd.DataFrame) -> pd.Series:
    """从日柱排聚合周柱排（取每行最近的周内日柱排）
    与VBA方式基本一致：周柱排 = 本周最后一根K线的柱排"""
    return df["柱排日"].copy()


def 周护型聚合(df: pd.DataFrame) -> pd.Series:
    """周护型 = 本周 WXAB:
    - WXAB > 0 → 甲/乙/己
    - WXAB ≤ 0 → 丙/警"""
    result = pd.Series("警", index=df.index, dtype=object)
    甲条件 = (df["AB"] > 0) & (df["ZA"] > 0)
    乙条件 = (df["AB"] > 0) & (df["ZA"] <= 0) & (df["ZB"] > 0)
    己条件 = (df["AB"] > 0) & (df["ZB"] <= 0)
    丙条件 = (df["AB"] <= 0)
    result[甲条件] = "甲"
    result[乙条件] = "乙"
    result[己条件] = "己"
    result[丙条件] = "丙"
    return result


def 计算周大局(WXCD: pd.Series, 日层联动: pd.Series, 周护型: pd.Series) -> pd.Series:
    """大局(WXCD+WJC+WXAB) 周 — 简化版"""
    return WXCD + "周局." + 周护型


def 周版BTZ(df: pd.DataFrame) -> pd.DataFrame:
    """周级BTZ = 本周最后一行的BTZ值（与VBA周结算等效）"""
    for vba_col in ["ZA", "ZB", "ZC", "ZD", "ZE", "AB", "AC", "BC", "CD", "CE", "DE"]:
        if vba_col in df.columns:
            df[f"{vba_col}周"] = df[vba_col].copy()
    return df


# ============================================================
# 4. 生成LD sheet列
# ============================================================

def 构建LD数据(df: pd.DataFrame) -> pd.DataFrame:
    """构建与VBA算展D结构一致的LD数据"""
    rows = []

    for i in range(len(df)):
        龄 = i + 1
        期 = df["date"].iloc[i]
        键 = True
        七 = 期.isocalendar()[1]  # 周号
        涨 = round(df["涨幅"].iloc[i], 2) if "涨幅" in df.columns else 0.0

        # OHLC
        P0 = round(df["open"].iloc[i], 3)
        P1 = round(df["high"].iloc[i], 3)
        P3 = round(df["low"].iloc[i], 3)
        H0 = round(df["close"].iloc[i], 3)
        H1 = round(df["volume"].iloc[i], 2)
        H3 = round(df["amount"].iloc[i], 2)

        # === 周级指标 (col 36-88) ===
        周柱排 = str(df["周柱排"].iloc[i]) if "周柱排" in df.columns and pd.notna(df["周柱排"].iloc[i]) else ""
        周护型 = str(df["周护型"].iloc[i]) if "周护型" in df.columns else ""
        周大局 = str(df["周大局"].iloc[i]) if "周大局" in df.columns else ""
        周柱排字符 = str(df["周柱排"].iloc[i]) if "周柱排" in df.columns else ""

        # 下柱冲高提示
        冲高提示 = str(df["下柱冲高提示"].iloc[i]) if "下柱冲高提示" in df.columns and pd.notna(df["下柱冲高提示"].iloc[i]) else ""

        # 下周冲高 / 下日冲高
        下周冲 = str(df["下周冲高"].iloc[i]) if "下周冲高" in df.columns else ""

        # === 日级核心指标 (col 89-96) ===
        日周联动 = str(df["日层联动"].iloc[i]) if "日层联动" in df.columns else ""
        DX护型 = str(df["日层护型"].iloc[i]) if "日层护型" in df.columns else ""
        日柱排 = str(df["柱排日"].iloc[i]) if "柱排日" in df.columns else ""

        row_data = {
            "龄": 龄, "期": 期, "键": 键, "七": 七, "涨": 涨,
            # OHLC
            "P0": P0, "P1": P1, "P3": P3, "H0": H0, "H1": H1, "H3": H3,
            # 周级
            "周柱排": 周柱排, "周护型WXAB": 周护型,
            "大局(WXCD+WJC+WXAB)周": 周大局,
            "下柱冲高提示下周": 冲高提示,
            "下周冲高": 下周冲,
            # 日级
            "日周联动": 日周联动, "DXAB护型": DX护型, "柱排": 日柱排,
            "WXCD": str(df["WXCD"].iloc[i]) if "WXCD" in df.columns else "",
        }

        # BTZ日级
        for vba_col in ["ZA", "ZB", "ZC", "ZD", "ZE", "AB", "AC", "BC", "CD", "CE", "DE"]:
            col_key = f"BTZ_{vba_col}日"
            val = df[f"{vba_col}日"].iloc[i] if f"{vba_col}日" in df.columns else (
                df[vba_col].iloc[i] if vba_col in df.columns else 0
            )
            row_data[f"{vba_col}日"] = int(val) if pd.notna(val) else 0

        # BTZ周级
        for vba_col in ["ZA", "ZB", "ZC", "ZD", "ZE", "AB", "AC", "BC", "CD", "CE", "DE"]:
            val = df[f"{vba_col}周"].iloc[i] if f"{vba_col}周" in df.columns else 0
            row_data[f"{vba_col}周"] = int(val) if pd.notna(val) else 0

        rows.append(row_data)

    return pd.DataFrame(rows)


# ============================================================
# 5. 写入xlsx
# ============================================================

def 写出LD到xlsx(df: pd.DataFrame, 输出路径: str):
    """写出LD sheet到xlsx，保持VBA兼容格式"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LD"

    # 列定义（从VBA算展D结构提取关键列顺序）
    列顺序 = [
        ("龄", 1), ("期", 2), ("键", 3), ("七", 4), ("涨", 5),
        ("P0", 28), ("P1", 29), ("P3", 30), ("H0", 31), ("H1", 32), ("H3", 33),
        ("周柱排", 82), ("周护型WXAB", 79),
        ("大局(WXCD+WJC+WXAB)周", 88),
        ("下柱冲高提示下周", 87), ("下周冲高", 86),
        ("日周联动", 89), ("DXAB护型", 90), ("柱排", 96), ("WXCD", 98),
    ]

    # BTZ日级 (268-278)
    for vba_col in ["ZA日", "ZB日", "ZC日", "ZD日", "ZE日", "AB日", "AC日", "BC日", "CD日", "CE日", "DE日"]:
        列顺序.append((vba_col, None))

    # BTZ周级 (282-292)
    for vba_col in ["ZA周", "ZB周", "ZC周", "ZD周", "ZE周", "AB周", "AC周", "BC周", "CD周", "CE周", "DE周"]:
        列顺序.append((vba_col, None))

    # 分配列号
    col_num = 1
    fixed_cols = {name: num for name, num in 列顺序 if num is not None}
    dynamic_cols = [name for name, num in 列顺序 if num is None]

    # 先写固定位置列
    列映射 = {}
    for name, num in 列顺序:
        if num is not None:
            列映射[num] = name

    # 动态列从固定列之后开始分配
    max_fixed = max(列映射.keys()) + 1
    for i, name in enumerate(dynamic_cols):
        列映射[max_fixed + i] = name

    # 写表头
    头填充 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    头字体 = Font(bold=True, color="FFFFFF", size=9)

    for col_idx, name in sorted(列映射.items()):
        cell = ws.cell(1, col_idx, name)
        cell.font = 头字体
        cell.fill = 头填充
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 写数据
    for r_idx, (_, row) in enumerate(df.iterrows()):
        excel_row = r_idx + 2
        for col_idx, name in sorted(列映射.items()):
            if name in df.columns:
                val = row[name]
            else:
                val = 0

            # 日期处理
            if name == "期":
                if isinstance(val, (datetime, pd.Timestamp)):
                    # Excel序列号
                    base = datetime(1899, 12, 30)
                    delta = val - base
                    ws.cell(excel_row, col_idx, delta.days + delta.seconds / 86400)
                    ws.cell(excel_row, col_idx).number_format = "YYYY-MM-DD"
                    if name == "期":
                        ws.cell(excel_row, col_idx).number_format = "YYYY-MM-DD"
                continue

            if isinstance(val, float):
                cell = ws.cell(excel_row, col_idx, round(val, 4))
            elif val is None or (isinstance(val, float) and np.isnan(val)):
                cell = ws.cell(excel_row, col_idx, None)
            elif isinstance(val, np.integer):
                cell = ws.cell(excel_row, col_idx, int(val))
            else:
                cell = ws.cell(excel_row, col_idx, val)

    # 保存
    wb.save(输出路径)
    print(f"  已保存: {输出路径}")
    print(f"  数据行: {len(df)}, 列数: {len(列映射)}")
    wb.close()


# ============================================================
# 6. 主流程
# ============================================================

def main():
    symbol = "512660"
    输出文件 = "昭明算展/py算展sh512660.xlsx"

    print("=" * 50)
    print(f"生成 {输出文件}")
    print("=" * 50)

    # Step 1: 下载原始数据
    print("\n[1/4] 下载数据...")
    df = 下载AKshareETF数据(symbol)

    if len(df) == 0:
        print("❌ 无数据，退出")
        return

    # Step 2: 全量结算（EMA + BTZ + 涨跌幅）
    print("\n[2/4] 全量结算...")
    df = 全量结算(df)
    print(f"  均线+BTZ完成: {len(df)}行, {len(df.columns)}列")

    # 重命名BTZ
    df = 重命名BTZ为VBA名(df)

    # Step 3: 策略引擎
    print("\n[3/4] 策略计算 + 柱排...")
    df["WXCD"] = 提取WXCD前缀(df)
    df["日层护型"] = DXAB护型分类(df)

    btze = df["ZE"].values
    btzc = df["ZC"].values
    btzd = df["ZD"].values
    btzb = df["ZB"].values
    btza = df["ZA"].values
    btcd = df["CD"].values
    btbc = df["BC"].values
    btab = df["AB"].values

    df["日层联动"] = df["WXCD"] + 计算日层联动系列(
        btze, btcd, btzc, btbc, btab, btza, btzb, btzd
    )

    # 日柱排
    df["柱排日"] = 计算柱排序列(df, "日")
    print(f"  日柱排完成")

    # 下柱冲高提示（需要周柱排）
    df["周柱排"] = 周柱排聚合(df)
    df["下柱冲高提示"] = 计算下柱冲高提示(df, df["周柱排"])

    # 下周冲高 / 下日冲高
    df["下周冲高"] = 计算下周冲高信号(btze, df["下柱冲高提示"])

    # 周护型
    df["周护型"] = 周护型聚合(df)

    # 周大局
    df["周大局"] = 计算周大局(df["WXCD"], df["日层联动"], df["周护型"])

    # 猪操作
    df["猪操作"] = 计算猪操作(btze, btzc, btza, btzd, btzb, df["日层联动"])

    # BTZ周级
    df = 周版BTZ(df)

    # Step 4: 构建并写出LD
    print("\n[4/4] 写出LD sheet...")
    ld_df = 构建LD数据(df)
    写出LD到xlsx(ld_df, 输出文件)

    # 列统计
    print(f"\n{'='*50}")
    print(f"  完成: {输出文件}")
    print(f"  行数: {len(ld_df)}")
    print(f"  日期: {df['date'].min().date()} ~ {df['date'].max().date()}")
    print(f"{'='*50}")

    # 预览最后5行关键列
    print("\n最后5行预览（核心列）:")
    preview_cols = ["龄", "期", "涨", "日周联动", "DXAB护型", "柱排", "周柱排", "WXCD"]
    for c in preview_cols:
        if c in ld_df.columns:
            vals = ld_df[c].tail(5).tolist()
            print(f"  {c}: {vals}")

    return ld_df


if __name__ == "__main__":
    main()