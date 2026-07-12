"""
数据加载引擎 — 读原始OHLCV → 算EMA均线 → 算BTZ交叉 → 等价于VBA 组结算数组

均线参数（来自 IQQQ跨码_D据擎5结算.bas）:
  JA: EMA(alpha=2/6)  ≈ 5期
  JB: EMA(alpha=2/13) ≈ 12期
  JC: EMA(alpha=2/27) ≈ 26期
  JD: EMA(alpha=2/61) ≈ 60期
  JE: EMA(alpha=2/121)≈ 120期

BTZ逻辑（来自 STBASE结算工具_衍生交叉天数）:
  正数 = 在均线上方持续N天；负数 = 在均线下方持续|N|天
  +/-1 = 刚交叉
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Optional, Dict


# ============================================================
# 1. 原始数据加载
# ============================================================

def 读取CSV(path: str, 编码: str = "gbk") -> pd.DataFrame:
    """读AKshare格式CSV：日期,开盘,最高,最低,收盘,成交量,成交额"""
    df = pd.read_csv(path, encoding=编码)
    # 统一列名
    列映射 = {
        "日期": "date", "开盘": "open", "最高": "high",
        "最低": "low", "收盘": "close", "成交量": "volume",
        "成交额": "amount", "date": "date", "open": "open",
        "high": "high", "low": "low", "close": "close",
        "volume": "volume", "amount": "amount",
    }
    df.rename(columns=列映射, inplace=True)
    # 只保留需要的列
    df = df[["date", "open", "high", "low", "close", "volume", "amount"]]
    df["date"] = pd.to_datetime(df["date"])
    df.sort_values("date", inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


def 读取ZdataCSV(代码: str, 数据目录: str = r"D:\zdata\data中股") -> pd.DataFrame:
    """从zdata本地CSV仓库读取OHLCV数据（无表头）

    格式A（7列，ETF/指数）: id,date,open,high,low,close,volume
    格式B（9列，个股）:     id,date,open,high,low,close,volume,total_shares,turnover_rate

    返回统一列名: date,open,high,low,close,volume,amount
    """
    import glob
    路径 = Path(数据目录) / f"{代码}.csv"
    if not 路径.exists():
        raise FileNotFoundError(f"Zdata CSV不存在: {路径}")

    df = pd.read_csv(路径, header=None)
    n_cols = df.shape[1]
    if n_cols == 7:
        df.columns = ["id", "date", "open", "high", "low", "close", "volume"]
        df["amount"] = 0.0
    elif n_cols == 9:
        df.columns = ["id", "date", "open", "high", "low", "close", "volume", "total_shares", "turnover_rate"]
        # 估算成交额 = volume * close 的近似值，或用 total_shares * close * turnover_rate
        df["amount"] = df["volume"] * df["close"]
    else:
        raise ValueError(f"zdata CSV列数异常: {n_cols} (预期7或9)")

    df = df[["date", "open", "high", "low", "close", "volume", "amount"]]
    df["date"] = pd.to_datetime(df["date"])
    df.sort_values("date", inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


# ============================================================
# 2. EMA均线计算（与VBA IQQQ跨码_D据擎5结算 完全一致）
# ============================================================

def 计算VBA均线(df: pd.DataFrame) -> pd.DataFrame:
    """与VBA结算引擎均线基础迭代完全一致

    VBA算法:
      1. JZ = Round((1/3)*prev_JZ + (2/3)*close, 3)  — alpha=2/3 ≈ 2期EMA
      2. JA = Round((4/6)*prev_JA + (2/6)*JZ, 3)      — alpha=2/6 ≈ 5期
      3. JB = Round((11/13)*prev_JB + (2/13)*JZ, 3)   — alpha=2/13 ≈ 12期
      4. JC = Round((25/27)*prev_JC + (2/27)*JZ, 3)   — alpha=2/27 ≈ 26期
      5. JD = Round((59/61)*prev_JD + (2/61)*JZ, 3)   — alpha=2/61 ≈ 60期
      6. JE = Round((119/121)*prev_JE + (2/121)*JZ, 3) — alpha=2/121 ≈ 120期

      初始化: 所有均线 = 第一日收盘价
      关键: JA~JE 是对 JZ 做EMA, 不是直接对 close!
    """
    close = df["close"].values
    n = len(close)
    df["JZ"] = np.zeros(n)
    df["JA"] = np.zeros(n)
    df["JB"] = np.zeros(n)
    df["JC"] = np.zeros(n)
    df["JD"] = np.zeros(n)
    df["JE"] = np.zeros(n)

    # 首日：全部初始化为收盘价
    df["JZ"].iloc[0] = round(close[0], 3)
    df["JA"].iloc[0] = round(close[0], 3)
    df["JB"].iloc[0] = round(close[0], 3)
    df["JC"].iloc[0] = round(close[0], 3)
    df["JD"].iloc[0] = round(close[0], 3)
    df["JE"].iloc[0] = round(close[0], 3)

    JZ_prev = df["JZ"].iloc[0]
    JA_prev = df["JA"].iloc[0]
    JB_prev = df["JB"].iloc[0]
    JC_prev = df["JC"].iloc[0]
    JD_prev = df["JD"].iloc[0]
    JE_prev = df["JE"].iloc[0]

    for i in range(1, n):
        # JZ = EMA(close, alpha=2/3)
        JZ_cur = round((1/3) * JZ_prev + (2/3) * close[i], 3)
        df["JZ"].iloc[i] = JZ_cur

        # JA~JE 都是对 JZ 做EMA（不是对close！）
        df["JA"].iloc[i] = round((4/6) * JA_prev + (2/6) * JZ_cur, 3)
        df["JB"].iloc[i] = round((11/13) * JB_prev + (2/13) * JZ_cur, 3)
        df["JC"].iloc[i] = round((25/27) * JC_prev + (2/27) * JZ_cur, 3)
        df["JD"].iloc[i] = round((59/61) * JD_prev + (2/61) * JZ_cur, 3)
        df["JE"].iloc[i] = round((119/121) * JE_prev + (2/121) * JZ_cur, 3)

        JZ_prev = JZ_cur
        JA_prev = df["JA"].iloc[i]
        JB_prev = df["JB"].iloc[i]
        JC_prev = df["JC"].iloc[i]
        JD_prev = df["JD"].iloc[i]
        JE_prev = df["JE"].iloc[i]

    return df


# 向后兼容别名
计算均线 = 计算VBA均线


# ============================================================
# 3. BTZ交叉天数计算
# ============================================================

def 计算BTZ(快线: np.ndarray, 慢线: np.ndarray) -> np.ndarray:
    """与VBA STBASE结算工具_衍生交叉天数 完全一致"""
    result = np.zeros(len(快线), dtype=int)
    for i in range(1, len(快线)):
        前值 = result[i-1]
        if 前值 > 0:
            if 快线[i] >= 慢线[i]:
                result[i] = 前值 + 1
            else:
                result[i] = -1
        else:
            if 快线[i] < 慢线[i]:
                result[i] = 前值 - 1
            else:
                result[i] = 1
    return result


def 计算全部BTZ(df: pd.DataFrame) -> pd.DataFrame:
    """计算所有BTZ列（与VBA结算引擎完全一致）

    VBA算法(IQQQ跨码_D据擎5结算.bas:308+):
      价格线 = JZ (不是close! JZ = EMA(close, 2/3))
      均线 = JA/JB/JC/JD/JE (对JZ做EMA)
    """
    JZ = df["JZ"].values if "JZ" in df.columns else df["close"].values
    JA = df["JA"].values
    JB = df["JB"].values
    JC = df["JC"].values
    JD = df["JD"].values
    JE = df["JE"].values

    # BTZ(JZ, JA~JE) — VBA: BTZA=BTZ(JZ,JA), BTZB=BTZ(JZ,JB), ...
    df["BTZA"] = 计算BTZ(JZ, JA)
    df["BTZB"] = 计算BTZ(JZ, JB)
    df["BTZC"] = 计算BTZ(JZ, JC)
    df["BTZD"] = 计算BTZ(JZ, JD)
    df["BTZE"] = 计算BTZ(JZ, JE)

    # 均线 vs 均线（排序关系）
    df["BTAB"] = 计算BTZ(JA, JB)  # DJA vs DJB
    df["BTAC"] = 计算BTZ(JA, JC)  # DJA vs DJC
    df["BTBC"] = 计算BTZ(JB, JC)  # DJB vs DJC
    df["BTAD"] = 计算BTZ(JA, JD)  # DJA vs DJD
    df["BTBD"] = 计算BTZ(JB, JD)  # DJB vs DJD
    df["BTCD"] = 计算BTZ(JC, JD)  # DJC vs DJD
    df["BTCE"] = 计算BTZ(JC, JE)  # DJC vs DJE
    df["BTDE"] = 计算BTZ(JD, JE)  # DJD vs DJE

    return df


# ============================================================
# 4. 涨跌幅计算
# ============================================================

def 计算涨跌幅(df: pd.DataFrame) -> pd.DataFrame:
    """与VBA一致: (今收/前收 - 1) * 100"""
    df["涨幅"] = df["close"].pct_change() * 100
    df["涨幅"].fillna(0, inplace=True)
    df["涨幅"] = df["涨幅"].round(2)
    return df


# ============================================================
# 5. 日线→周线/月线 聚合（与VBA STBASE取据引擎_工具结算数据转换 等效）
# ============================================================

def 聚合日线为周期(df: pd.DataFrame, 周期: str = "W") -> pd.DataFrame:
    """将日线OHLCV数据聚合为周线或月线

    参数:
        df: 每日数据DataFrame，必须含 date, open, high, low, close, volume, amount 列
        周期: "W"=周线, "M"=月线

    返回:
        聚合后的DataFrame（与VBA STBASE取据引擎_工具结算数据转换 逻辑一致）:
        - open  = 周期首日开盘
        - high  = 周期内最高
        - low   = 周期内最低
        - close = 周期末日收盘
        - volume = 周期总成交量
        - amount = 周期总成交额
        - date  = 周期末日日期
        - 龄    = 周期序号（从1开始）
        - 是末天 = True（仅保留每个周期的最后一行，与VBA同样行为）

    VBA算法对照:
        - 周: 周日为起点（Weekday date, vbMonday）= 7对应周日
          周起始 = date - (Weekday(date, vbMonday) - 1)
        - 月: 起始 = date - (Day(date) - 1)
        Python: 周一起始 = (date - timedelta(days=weekday-1)) ，周日为终点
        使用ISO week标准：周一=1 周日=7
        周起始 = 周一，周结束 = 周日
    """
    df = df.copy()
    df = df.sort_values("date").reset_index(drop=True)

    if 周期 == "W":
        # 周起始 = 当前周的周一，周结束 = 周日
        df["期类第几日"] = df["date"].apply(lambda d: d.isoweekday())  # 1=Mon ... 7=Sun
        df["期类期始"] = df["date"] - pd.to_timedelta(df["期类第几日"] - 1, unit="D")
        分组键 = df["期类期始"]
    elif 周期 == "M":
        df["期类第几日"] = df["date"].apply(lambda d: d.day)
        df["期类期始"] = df["date"] - pd.to_timedelta(df["期类第几日"] - 1, unit="D")
        分组键 = df["期类期始"]
    else:
        raise ValueError(f"不支持的周期: {周期}，仅支持 'W'(周) 或 'M'(月)")

    # 分组聚合（与VBA逻辑一致：首开/最高/最低/末收/总量）
    聚合结果 = df.groupby(分组键, sort=True).agg(
        date=("date", "last"),
        open=("open", "first"),
        high=("high", "max"),
        low=("low", "min"),
        close=("close", "last"),
        volume=("volume", "sum"),
        amount=("amount", "sum"),
        期类第几日=("期类第几日", "first"),
        期类期始=("期类期始", "first"),
    ).reset_index(drop=True)

    # 龄 = 周期序号
    聚合结果.insert(0, "龄", range(1, len(聚合结果) + 1))
    聚合结果["是末天"] = True  # 每个周期仅保留最后一行
    聚合结果["周期"] = 周期

    return 聚合结果


# ============================================================
# 6. 主力入口
# ============================================================

def 全量结算(df: pd.DataFrame) -> pd.DataFrame:
    """从原始OHLCV到完整组结算等价DataFrame (VBA兼容版)"""
    df = df.copy()
    df = 计算VBA均线(df)
    df = 计算全部BTZ(df)
    df = 计算涨跌幅(df)
    return df


def 从CSV加载并结算(path: str) -> pd.DataFrame:
    """从CSV文件加载原始数据并完成全量结算"""
    df = 读取CSV(path)
    df = 全量结算(df)
    return df


def 从算展D加载(path: str) -> pd.DataFrame:
    """从已结算的算展D.xlsx加载（含已有BTZ值，用于验证）"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    # 找到LD sheet
    表名 = [s for s in wb.sheetnames if s.startswith("LD")][0]
    ws = wb[表名]

    # 读列头
    列头 = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val:
            列头[col] = str(val).strip()

    # 读数据
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        rows.append(row)

    wb.close()
    df = pd.DataFrame(rows)

    # 列名索引（从列头反查）
    列索引 = {v: k for k, v in 列头.items()}
    print(f"  列数: {len(列索引)}, 数据行: {len(df)}")
    return df, 列头, 列索引


# ============================================================
# 测试入口
# ============================================================
if __name__ == "__main__":
    # 测试：从算展D加载验证列位置
    df, headers, col_idx = 从算展D加载("算展D.sz300304.xlsx")
    print(f"  关键列: BTZE={col_idx.get('ZE 日','?')}, BTZA={col_idx.get('ZA 日','?')}")