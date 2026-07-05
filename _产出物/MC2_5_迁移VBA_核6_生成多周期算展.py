"""
生成多周期 py算展*.xlsx — 日/周/月 三周期全量结算

流程（对每个周期）:
  1. AKshare下载日线OHLCV
  2. 按周期聚合（日=保留原样 / 周=周线聚合 / 月=月线聚合）
  3. 全量结算（EMA均线 + BTZ交叉 + 涨跌幅）
  4. 策略引擎（WXCD前缀 + 日层护型 + 日层联动 + 猪操作）
  5. 柱排引擎（柱排 + 下柱冲高提示）
  6. 写入LD sheet → 昭明算展/py算展{code}_{周期}.xlsx

与VBA版算展D列结构对齐:
  基本列: 龄, 期, 键, 七(周号), 涨
  OHLC:  P0~P3, H0~H3
  周级:  柱排周, 周护型, 大局, 冲高提示
  日级:  日周联动, DXAB护型, 柱排, WXCD
  BTZ:   ZA~DE (日级), ZA~DE (周级)
"""

import sys, os
sys.path.insert(0, os.path.dirname(__file__))

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path

from MC2_5_迁移VBA_核1_数据加载引擎 import 全量结算, 聚合日线为周期
from MC2_5_迁移VBA_核2_策略计算引擎 import 提取WXCD前缀, DXAB护型分类, 计算日层联动系列, 计算猪操作
from MC2_5_迁移VBA_核3_柱排解析引擎 import 计算柱排序列

import warnings
warnings.filterwarnings("ignore")

# BTZ列名映射：Python名 → VBA名
BTZ_VBA名 = {
    "BTZA": "ZA", "BTZB": "ZB", "BTZC": "ZC",
    "BTZD": "ZD", "BTZE": "ZE",
    "BTAB": "AB", "BTAC": "AC", "BTBC": "BC",
    "BTAD": "AD", "BTBD": "BD", "BTCD": "CD",
    "BTCE": "CE", "BTDE": "DE",
}


def 下载数据(symbol: str, 起始日: str = "20160101", 结束日: str = "20260630"):
    """下载ETF日线数据"""
    import akshare as ak
    df = ak.fund_etf_hist_em(
        symbol=symbol, period="daily",
        start_date=起始日, end_date=结束日, adjust="qfq"
    )
    列映射 = {"日期": "date", "开盘": "open", "最高": "high",
               "最低": "low", "收盘": "close", "成交量": "volume", "成交额": "amount"}
    df.rename(columns=列映射, inplace=True)
    df = df[["date", "open", "high", "low", "close", "volume", "amount"]]
    df["date"] = pd.to_datetime(df["date"])
    df.sort_values("date", inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


def 重命名BTZ(df):
    """Python BTZ列名 → VBA名 (ZA/ZB/AB等)"""
    for py_col, vba_col in BTZ_VBA名.items():
        if py_col in df.columns:
            df[vba_col] = df[py_col].astype(int)
    return df


def 计算策略列(df: pd.DataFrame) -> pd.DataFrame:
    """对单周期DataFrame执行策略计算"""
    df["WXCD"] = 提取WXCD前缀(df)
    df["护型DXAB"] = DXAB护型分类(df)

    btze = df["ZE"].values
    btzc = df["ZC"].values
    btzd = df["ZD"].values
    btzb = df["ZB"].values
    btza = df["ZA"].values
    btcd = df["CD"].values
    btbc = df["BC"].values
    btab = df["AB"].values

    df["日层联动"] = df["WXCD"] + 计算日层联动系列(
        btze, btcd, btzc, btbc, btab, btza, btzb, btzd
    )
    # 柱排用"日"标记（实际按周期数据计算）
    df["柱排"] = 计算柱排序列(df, "日")
    # 猪操作
    df["猪操作"] = 计算猪操作(btze, btzc, btza, btzd, btzb, df["日层联动"])
    return df


def 构建LD数据(df: pd.DataFrame, 周期名: str) -> pd.DataFrame:
    """构建与VBA算展D结构一致的LD数据帧"""
    rows = []
    for i in range(len(df)):
        期 = df["date"].iloc[i]
        龄 = i + 1
        键 = True
        # 周号（不按实际周计算，按数据位置）
        七 = 期.isocalendar()[1] if hasattr(期, 'isocalendar') else (i % 52 + 1)
        涨 = round(df["涨幅"].iloc[i], 2) if "涨幅" in df.columns else 0.0

        row = {
            "龄": 龄, "期": 期, "键": 键, "七": 七, "涨": 涨,
            "P0": round(df["open"].iloc[i], 3),
            "P1": round(df["high"].iloc[i], 3),
            "P3": round(df["low"].iloc[i], 3),
            "H0": round(df["close"].iloc[i], 3),
            "H1": round(df["volume"].iloc[i], 2),
            "H3": round(df["amount"].iloc[i], 2),
            # 策略列
            "柱排": str(df["柱排"].iloc[i]) if "柱排" in df.columns else "",
            "日周联动": str(df["日层联动"].iloc[i]) if "日层联动" in df.columns else "",
            "DXAB护型": str(df["护型DXAB"].iloc[i]) if "护型DXAB" in df.columns else "",
            "WXCD": str(df["WXCD"].iloc[i]) if "WXCD" in df.columns else "",
            "猪操作": str(df["猪操作"].iloc[i]) if "猪操作" in df.columns else "",
        }

        # BTZ列
        for vba_col in ["ZA", "ZB", "ZC", "ZD", "ZE", "AB", "AC", "BC", "CD", "CE", "DE"]:
            val = df[vba_col].iloc[i] if vba_col in df.columns else 0
            row[f"{vba_col}"] = int(val) if pd.notna(val) else 0

        rows.append(row)
    return pd.DataFrame(rows)


def 写出xlsx(ld_df: pd.DataFrame, 输出路径: str, 周期名: str):
    """写出LD sheet到xlsx"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"LD{周期名}"

    # 列定义（位置尽量靠近VBA标准）
    all_cols = [
        "龄", "期", "键", "七", "涨",
        "P0", "P1", "P3", "H0", "H1", "H3",
    ]
    # 策略列（指定位置）
    策略列 = [
        ("柱排", 6), ("日周联动", 7), ("DXAB护型", 8),
        ("WXCD", 9), ("猪操作", 10),
    ]
    # BTZ列（在后）
    btz_cols = [f"{v}日" for v in ["ZA", "ZB", "ZC", "ZD", "ZE", "AB", "AC", "BC", "CD", "CE", "DE"]]
    btz_cols += [f"{v}周" for v in ["ZA", "ZB", "ZC", "ZD", "ZE", "AB", "AC", "BC", "CD", "CE", "DE"]]

    # 映射最终列号
    col_num = 1
    列映射 = {}
    for name in all_cols:
        列映射[col_num] = name
        col_num += 1

    # BTZ放在策略列之前？不用，策略列固定位置
    策略列映射 = {}
    for name, pos in 策略列:
        策略列映射[pos] = name

    # 在数据中查找策略列并设置
    final_cols = {}
    for cn in range(1, max(列映射.keys()) + 1):
        if cn in 列映射:
            final_cols[cn] = 列映射[cn]

    # 找未占用的列号放策略列和BTZ
    used = set(final_cols.keys())
    next_col = max(used) + 1
    for name, pos in 策略列:
        if pos not in used:
            final_cols[pos] = name
            used.add(pos)
        else:
            while next_col in used:
                next_col += 1
            final_cols[next_col] = name
            used.add(next_col)

    for col in btz_cols:
        while next_col in used:
            next_col += 1
        final_cols[next_col] = col
        used.add(next_col)

    # 更简单的做法：按顺序写所有列
    列顺序 = all_cols + [c[0] for c in 策略列] + btz_cols

    # 写表头
    头填充 = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    头字体 = Font(bold=True, color="FFFFFF", size=9)

    for col_idx, name in enumerate(列顺序, 1):
        cell = ws.cell(1, col_idx, name)
        cell.font = 头字体
        cell.fill = 头填充
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 写数据
    for r_idx, (_, row) in enumerate(ld_df.iterrows()):
        excel_row = r_idx + 2
        for col_idx, name in enumerate(列顺序, 1):
            if name in ld_df.columns:
                val = row[name]
            else:
                val = 0 if "_日" not in name and "_周" not in name else 0

            if name == "期":
                if isinstance(val, (datetime, pd.Timestamp)):
                    base = datetime(1899, 12, 30)
                    delta = val - base
                    cell = ws.cell(excel_row, col_idx, delta.days + delta.seconds / 86400)
                    cell.number_format = "YYYY-MM-DD"
                    continue
            elif isinstance(val, float):
                cell = ws.cell(excel_row, col_idx, round(val, 4))
            elif val is None or (isinstance(val, float) and np.isnan(val)):
                cell = ws.cell(excel_row, col_idx, None)
            elif isinstance(val, np.integer):
                cell = ws.cell(excel_row, col_idx, int(val))
            else:
                cell = ws.cell(excel_row, col_idx, val)

    wb.save(输出路径)
    print(f"  已保存: {输出路径}  ({len(ld_df)}行, {len(列顺序)}列)")
    wb.close()


def 生成单周期(df_日线: pd.DataFrame, 周期: str, 周期名: str) -> pd.DataFrame:
    """对日线数据做周期聚合+全量结算+策略计算，返回LD格式DataFrame"""
    print(f"  [{周期名}] 聚合日线...")
    if 周期 == "D":
        df = df_日线.copy()
    else:
        df = 聚合日线为周期(df_日线, 周期)
    print(f"    {周期名}数据: {len(df)}行")

    print(f"  [{周期名}] 全量结算...")
    df = 全量结算(df)
    df = 重命名BTZ(df)

    print(f"  [{周期名}] 策略计算+柱排...")
    df = 计算策略列(df)

    print(f"  [{周期名}] 构建LD...")
    ld = 构建LD数据(df, 周期名)

    return ld


def 生成全周期(symbol: str, 输出目录: str = "昭明算展"):
    """生成日/周/月三周期算展文件"""
    路径 = Path(输出目录)
    路径.mkdir(exist_ok=True)

    print("=" * 55)
    print(f"  生成多周期算展: {symbol}")
    print("=" * 55)

    # 1. 下载日线
    print("\n[1] 下载日线数据...")
    df_日线 = 下载数据(symbol)
    print(f"    日线: {len(df_日线)}行 ({df_日线['date'].min().date()} ~ {df_日线['date'].max().date()})")

    # 2. 生成各周期
    结果 = {}
    for 周期, 周期名 in [("D", "日"), ("W", "周"), ("M", "月")]:
        print(f"\n[{周期名}] 处理中...")
        ld = 生成单周期(df_日线, 周期, 周期名)
        输出文件 = 路径 / f"py算展{symbol}_{周期}.xlsx"
        写出xlsx(ld, str(输出文件), 周期名)
        结果[周期] = ld

    print(f"\n{'=' * 55}")
    print(f"  全部完成: {symbol}")
    print(f"  文件:")
    print(f"    日线: py算展{symbol}_D.xlsx  ({len(结果['D'])}行)")
    print(f"    周线: py算展{symbol}_W.xlsx  ({len(结果['W'])}行)")
    print(f"    月线: py算展{symbol}_M.xlsx  ({len(结果['M'])}行)")
    print(f"{'=' * 55}")
    return 结果


def main():
    """批量生成多只股票的周期算展"""
    symbols = [
        "512660",  # 光伏ETF (主力测试)
    ]
    for symbol in symbols:
        try:
            生成全周期(symbol)
        except Exception as e:
            import traceback
            traceback.print_exc()
            print(f"\n  ❌ {symbol}: {e}")


if __name__ == "__main__":
    main()